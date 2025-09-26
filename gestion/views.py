from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.urls import reverse_lazy
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
import json
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.mail import send_mass_mail
from django.template.loader import render_to_string
from django.core.mail import send_mail
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from django.views.decorators.csrf import csrf_protect
from django.conf import settings
import os
import shutil
import logging
import time

from .models import Adherent, Section, Competence, GroupeCompetence, Seance, Evaluation, LienEvaluation, Palanquee, Lieu, LienInscriptionSeance, InscriptionSeance, Exercice
from .forms import AdherentForm, SectionForm, CompetenceForm, GroupeCompetenceForm, SeanceForm, EvaluationBulkForm, PalanqueeForm, NonAdherentInscriptionForm, AdherentPublicForm, ExerciceForm, AdminInscriptionSeanceForm, AffectationSectionMasseForm, CommunicationSeanceForm, CommunicationAdherentsForm
from .utils import envoyer_lien_evaluation, envoyer_lien_evaluation_avec_cc
from .models import PalanqueeEleve
from gestion.models import EvaluationExercice, GroupeCompetence, Competence, Exercice, Adherent
from django.contrib.admin.views.decorators import staff_member_required
from django.core.mail import EmailMessage
import tempfile
from gestion.palanquee_views import generer_fiche_palanquee_pdf
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from gestion.utils import get_signature_html
from django.core.mail import EmailMultiAlternatives
from email.mime.image import MIMEImage
from django.contrib.auth.models import User, Group
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.contrib.sites.shortcuts import get_current_site
from .utils import eleve_only, encadrant_only, admin_only, group_required
from django.utils.decorators import method_decorator
from .models import ModeleMailSeance
from .models import Adherent, Section, ModeleMailAdherents, HistoriqueMailAdherents

# Vues d'accueil et de navigation
@group_required('admin')
def dashboard(request):
    """Tableau de bord principal"""
    context = {
        'total_adherents': Adherent.objects.filter(type_personne='adherent').count(),
        'total_seances': Seance.objects.count(),
        'total_palanquees': Palanquee.objects.count(),
        'seances_recentes': Seance.objects.all()[:5],
        'palanquees_recentes': Palanquee.objects.select_related('seance', 'section').prefetch_related('eleves')[:5],
        'adherents_eleves': Adherent.objects.filter(type_personne='adherent', statut='eleve').count(),
        'adherents_encadrants': Adherent.objects.filter(type_personne='adherent', statut='encadrant').count(),
    }
    # Ajout des alertes CACI
    from datetime import timedelta
    today = timezone.now().date()
    adherents_caci = Adherent.objects.exclude(niveau='debutant')
    context['adherents_sans_caci'] = adherents_caci.filter(caci_fichier__isnull=True) | adherents_caci.filter(caci_fichier='')
    context['adherents_caci_expire'] = adherents_caci.filter(date_delivrance_caci__isnull=False, date_delivrance_caci__lt=today - timedelta(days=365))
    context['adherents_caci_bientot'] = adherents_caci.filter(date_delivrance_caci__isnull=False, date_delivrance_caci__gte=today - timedelta(days=365), date_delivrance_caci__lte=today - timedelta(days=335))
    context['adherents_caci_non_valide'] = adherents_caci.filter(caci_valide=False)
    # Bloc dernières palanquées évaluées (par encadrant)
    # On récupère les palanquées qui ont au moins une évaluation_exercice
    palanquees_evaluees_ids = (
        EvaluationExercice.objects.values_list('palanquee_id', flat=True).distinct()
    )
    palanquees_evaluees = (
        Palanquee.objects.filter(id__in=palanquees_evaluees_ids, encadrant__isnull=False)
        .select_related('seance', 'encadrant')
        .prefetch_related('eleves')
        .order_by('-seance__date', '-id')
        .distinct()
    )[:10]
    context['dernieres_palanquees_evaluees'] = palanquees_evaluees
    # Lien vers la page toutes les évaluations (à créer)
    context['url_toutes_evaluations'] = '/evaluations/'
    return render(request, 'gestion/dashboard.html', context)

# Vues pour les adhérents
@method_decorator(group_required('admin'), name='dispatch')
class AdherentListView(LoginRequiredMixin, ListView):
    model = Adherent
    template_name = 'gestion/adherent_list.html'
    context_object_name = 'adherents'
    # paginate_by = 20  # Pagination supprimée pour afficher tous les adhérents
    
    def get_queryset(self):
        queryset = Adherent.objects.all()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(email__icontains=q)
            )
        sort = self.request.GET.get('sort', 'nom')
        order = self.request.GET.get('order', 'asc')
        valid_sorts = ['nom', 'prenom', 'email', 'niveau', 'statut', 'date_delivrance_caci', 'telephone']
        if sort not in valid_sorts:
            sort = 'nom'
        if order == 'desc':
            sort = '-' + sort
        queryset = queryset.order_by(sort, 'prenom' if sort.lstrip('-') == 'nom' else 'nom')
        # Ajout du statut CACI pour affichage
        from datetime import timedelta
        from django.utils import timezone
        today = timezone.now().date()
        adherents = list(queryset)
        for a in adherents:
            if a.date_delivrance_caci:
                expiration = a.date_delivrance_caci + timedelta(days=365)
                if expiration < today:
                    a.caci_status = 'expired'
                elif (expiration - today).days < 30:
                    a.caci_status = 'soon'
                else:
                    a.caci_status = ''
            else:
                a.caci_status = ''
        return adherents

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        adherents = context['adherents']
        context['adherents_adherents'] = [a for a in adherents if a.type_personne == 'adherent' and a.actif]
        context['adherents_non_adherents'] = [a for a in adherents if a.type_personne == 'non_adherent' and a.actif]
        context['adherents_non_adherents_desactives'] = [a for a in adherents if a.type_personne == 'non_adherent' and not a.actif]
        return context

@method_decorator(group_required('admin'), name='dispatch')
class AdherentDetailView(LoginRequiredMixin, DetailView):
    model = Adherent
    template_name = 'gestion/adherent_detail.html'
    context_object_name = 'adherent'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['today'] = timezone.now().date()
        # Ajout : séances où l'adhérent est inscrit
        from .models import InscriptionSeance
        seances_inscrit = (
            InscriptionSeance.objects
            .filter(personne=self.object)
            .select_related('seance')
            .order_by('-seance__date')
        )
        context['seances_inscrit'] = [ins.seance for ins in seances_inscrit]
        return context

@method_decorator(group_required('admin'), name='dispatch')
class AdherentCreateView(LoginRequiredMixin, CreateView):
    model = Adherent
    form_class = AdherentForm
    template_name = 'gestion/adherent_form.html'
    success_url = reverse_lazy('adherent_list')

@method_decorator(group_required('admin'), name='dispatch')
class AdherentUpdateView(LoginRequiredMixin, UpdateView):
    model = Adherent
    form_class = AdherentForm
    template_name = 'gestion/adherent_form.html'
    success_url = reverse_lazy('adherent_list')

@method_decorator(group_required('admin'), name='dispatch')
class AdherentDeleteView(LoginRequiredMixin, DeleteView):
    model = Adherent
    template_name = 'gestion/adherent_confirm_delete.html'
    success_url = reverse_lazy('adherent_list')

# Vues pour les élèves
@method_decorator(group_required('encadrant'), name='dispatch')
class EleveListView(LoginRequiredMixin, ListView):
    model = Adherent
    template_name = 'gestion/eleve_list.html'
    context_object_name = 'eleves'
    # paginate_by = 20  # Pagination supprimée pour afficher tous les élèves
    
    def get_queryset(self):
        queryset = Adherent.objects.filter(statut='eleve').order_by('nom', 'prenom')
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(email__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        eleves = context['eleves']
        context['eleves_adherents'] = [e for e in eleves if e.type_personne == 'adherent']
        context['eleves_non_adherents'] = [e for e in eleves if e.type_personne == 'non_adherent']
        return context

@method_decorator(group_required('admin'), name='dispatch')
class EncadrantListView(LoginRequiredMixin, ListView):
    model = Adherent
    template_name = 'gestion/encadrant_list.html'
    context_object_name = 'encadrants'
    # paginate_by = 20  # Pagination supprimée pour afficher tous les encadrants
    
    def get_queryset(self):
        queryset = Adherent.objects.filter(statut='encadrant').order_by('nom', 'prenom')
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(email__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        encadrants = context['encadrants']
        context['encadrants_adherents'] = [e for e in encadrants if e.type_personne == 'adherent']
        context['encadrants_non_adherents'] = [e for e in encadrants if e.type_personne == 'non_adherent']
        return context

# Vues pour les sections
@method_decorator(group_required('admin'), name='dispatch')
class SectionListView(LoginRequiredMixin, ListView):
    model = Section
    template_name = 'gestion/section_list.html'
    context_object_name = 'sections'

@method_decorator(group_required('admin'), name='dispatch')
class SectionDetailView(LoginRequiredMixin, DetailView):
    model = Section
    template_name = 'gestion/section_detail.html'
    context_object_name = 'section'

@method_decorator(group_required('admin'), name='dispatch')
class SectionCreateView(LoginRequiredMixin, CreateView):
    model = Section
    form_class = SectionForm
    template_name = 'gestion/section_form.html'
    success_url = reverse_lazy('section_list')

@method_decorator(group_required('admin'), name='dispatch')
class SectionUpdateView(LoginRequiredMixin, UpdateView):
    model = Section
    form_class = SectionForm
    template_name = 'gestion/section_form.html'
    success_url = reverse_lazy('section_list')

@method_decorator(group_required('admin'), name='dispatch')
class SectionDeleteView(LoginRequiredMixin, DeleteView):
    model = Section
    template_name = 'gestion/section_confirm_delete.html'
    success_url = reverse_lazy('section_list')

# Vues pour les compétences
@method_decorator(group_required('admin'), name='dispatch')
class CompetenceListView(LoginRequiredMixin, ListView):
    model = Competence
    template_name = 'gestion/competence_list.html'
    context_object_name = 'competences'
    
    def get_queryset(self):
        queryset = Competence.objects.select_related('section')
        section_id = self.request.GET.get('section')
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sections'] = Section.objects.all()
        return context

@method_decorator(group_required('admin'), name='dispatch')
class CompetenceCreateView(LoginRequiredMixin, CreateView):
    model = Competence
    form_class = CompetenceForm
    template_name = 'gestion/competence_form.html'
    success_url = reverse_lazy('competence_list')

@method_decorator(group_required('admin'), name='dispatch')
class CompetenceUpdateView(LoginRequiredMixin, UpdateView):
    model = Competence
    form_class = CompetenceForm
    template_name = 'gestion/competence_form.html'
    success_url = reverse_lazy('competence_list')

@method_decorator(group_required('admin'), name='dispatch')
class CompetenceDeleteView(LoginRequiredMixin, DeleteView):
    model = Competence
    template_name = 'gestion/competence_confirm_delete.html'
    success_url = reverse_lazy('competence_list')

@method_decorator(group_required('admin'), name='dispatch')
class CompetenceDetailView(LoginRequiredMixin, DetailView):
    model = Competence
    template_name = 'gestion/competence_detail.html'
    context_object_name = 'competence'

# Vues pour les groupes de compétences
@method_decorator(group_required('admin'), name='dispatch')
class GroupeCompetenceListView(LoginRequiredMixin, ListView):
    model = GroupeCompetence
    template_name = 'gestion/groupe_competence_list.html'
    context_object_name = 'groupes'

@method_decorator(group_required('admin'), name='dispatch')
class GroupeCompetenceCreateView(LoginRequiredMixin, CreateView):
    model = GroupeCompetence
    form_class = GroupeCompetenceForm
    template_name = 'gestion/groupe_competence_form.html'
    success_url = reverse_lazy('groupe_competence_list')

@method_decorator(group_required('admin'), name='dispatch')
class GroupeCompetenceUpdateView(LoginRequiredMixin, UpdateView):
    model = GroupeCompetence
    form_class = GroupeCompetenceForm
    template_name = 'gestion/groupe_competence_form.html'
    success_url = reverse_lazy('groupe_competence_list')

@method_decorator(group_required('admin'), name='dispatch')
class GroupeCompetenceDeleteView(LoginRequiredMixin, DeleteView):
    model = GroupeCompetence
    template_name = 'gestion/groupe_competence_confirm_delete.html'
    success_url = reverse_lazy('groupe_competence_list')

@method_decorator(group_required('admin'), name='dispatch')
class GroupeCompetenceDetailView(LoginRequiredMixin, DetailView):
    model = GroupeCompetence
    template_name = 'gestion/groupe_competence_detail.html'
    context_object_name = 'groupe'

# Vues pour les séances
@method_decorator(group_required('admin'), name='dispatch')
class SeanceListView(LoginRequiredMixin, ListView):
    model = Seance
    template_name = 'gestion/seance_list.html'
    context_object_name = 'seances'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Seance.objects.prefetch_related('palanques')
        date_debut = self.request.GET.get('date_debut')
        date_fin = self.request.GET.get('date_fin')
        
        if date_debut:
            queryset = queryset.filter(date__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(date__lte=date_fin)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

@method_decorator(group_required('admin'), name='dispatch')
class SeanceDetailView(LoginRequiredMixin, DetailView):
    model = Seance
    template_name = 'gestion/seance_detail.html'
    context_object_name = 'seance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        seance = self.get_object()
        palanquees_qs = seance.palanques.select_related('section', 'encadrant').prefetch_related('eleves', 'competences')
        context['palanquees'] = palanquees_qs
        # Ajout pour affichage des inscrits et du covoiturage
        inscriptions = seance.inscriptions.select_related('personne').all()
        inscrits_encadrants = [i for i in inscriptions if i.personne.statut == 'encadrant']
        inscrits_eleves = [i for i in inscriptions if i.personne.statut == 'eleve']
        inscrits_non_adherents = [i for i in inscriptions if i.personne.type_personne == 'non_adherent']
        context['inscrits_encadrants'] = inscrits_encadrants
        context['inscrits_eleves'] = inscrits_eleves
        context['inscrits_non_adherents'] = inscrits_non_adherents
        context['nb_total_inscrits'] = len(inscriptions)
        context['nb_encadrants'] = len(inscrits_encadrants)
        context['nb_eleves'] = len(inscrits_eleves)
        context['covoiturage_propose'] = [i for i in inscriptions if i.covoiturage == 'propose']
        context['covoiturage_besoin'] = [i for i in inscriptions if i.covoiturage == 'besoin']
        return context

@method_decorator(group_required('admin'), name='dispatch')
class SeanceCreateView(LoginRequiredMixin, CreateView):
    model = Seance
    form_class = SeanceForm
    template_name = 'gestion/seance_form.html'
    success_url = reverse_lazy('seance_list')

@method_decorator(group_required('admin'), name='dispatch')
class SeanceUpdateView(LoginRequiredMixin, UpdateView):
    model = Seance
    form_class = SeanceForm
    template_name = 'gestion/seance_form.html'
    success_url = reverse_lazy('seance_list')

@method_decorator(group_required('admin'), name='dispatch')
class SeanceDeleteView(LoginRequiredMixin, DeleteView):
    model = Seance
    template_name = 'gestion/seance_confirm_delete.html'
    success_url = reverse_lazy('seance_list')

# Vues pour les évaluations
@login_required
def evaluation_detail(request, pk):
    """Voir le détail d'une évaluation"""
    evaluation = get_object_or_404(Evaluation, pk=pk)
    return render(request, 'gestion/evaluation_detail.html', {'evaluation': evaluation})

@login_required
def evaluation_update(request, pk):
    """Modifier une évaluation"""
    evaluation = get_object_or_404(Evaluation, pk=pk)
    
    if request.method == 'POST':
        form = EvaluationBulkForm(evaluation.palanquee, request.POST)
        if form.is_valid():
            # Sauvegarder l'évaluation
            note = form.cleaned_data.get(f'eval_{evaluation.eleve.id}_{evaluation.competence.id}')
            commentaire = form.cleaned_data.get(f'comment_{evaluation.eleve.id}_{evaluation.competence.id}')
            
            if note is not None:
                evaluation.note = note
                evaluation.commentaire = commentaire
                evaluation.save()
                messages.success(request, 'Évaluation mise à jour avec succès.')
                return redirect('palanquee_detail', pk=evaluation.palanquee.pk)
    else:
        form = EvaluationBulkForm(evaluation.palanquee)
    
    context = {
        'evaluation': evaluation,
        'form': form,
    }
    return render(request, 'gestion/evaluation_update.html', context)

@login_required
def evaluation_delete(request, pk):
    """Supprimer une évaluation"""
    evaluation = get_object_or_404(Evaluation, pk=pk)
    palanquee_pk = evaluation.palanquee.pk
    
    if request.method == 'POST':
        evaluation.delete()
        messages.success(request, 'Évaluation supprimée avec succès.')
        return redirect('palanquee_detail', pk=palanquee_pk)
    
    return render(request, 'gestion/evaluation_confirm_delete.html', {'evaluation': evaluation})

# Vues utilitaires
@login_required
def get_competences_section(request):
    """API pour récupérer les compétences d'une section (AJAX)"""
    section_id = request.GET.get('section_id')
    if section_id:
        competences = Competence.objects.filter(section_id=section_id).values('id', 'nom')
        return JsonResponse({'competences': list(competences)})
    return JsonResponse({'competences': []})

@login_required
def get_eleves_section(request):
    """API pour récupérer les élèves d'une section (AJAX)"""
    section_id = request.GET.get('section_id')
    if section_id:
        try:
            section = Section.objects.get(id=section_id)
            
            # D'abord, essayer de récupérer les élèves inscrits dans cette section
            eleves = Adherent.objects.filter(
                statut='eleve',
                sections=section
            ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
            
            # Si aucun élève n'est inscrit dans cette section, utiliser la logique de niveau
            if not eleves.exists():
                # Logique de filtrage des élèves selon la section (fallback)
                if section.nom == 'bapteme':
                    # Baptême : tous les élèves débutants
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau='debutant'
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                elif section.nom == 'prepa_niveau1':
                    # Prépa Niveau 1 : débutants et niveau 1
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau__in=['debutant', 'niveau1']
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                elif section.nom == 'prepa_niveau2':
                    # Prépa Niveau 2 : niveau 1 et niveau 2
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau__in=['niveau1', 'niveau2']
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                elif section.nom == 'prepa_niveau3':
                    # Prépa Niveau 3 : niveau 2 et niveau 3
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau__in=['niveau2', 'niveau3']
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                elif section.nom == 'prepa_niveau4':
                    # Prépa Niveau 4 : niveau 3
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau='niveau3'
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                elif section.nom == 'niveau3':
                    # Niveau 3 : niveau 3
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau='niveau3'
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                elif section.nom == 'niveau4':
                    # Niveau 4 : niveau 3 et plus
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau__in=['niveau3', 'initiateur1', 'initiateur2', 'moniteur_federal1', 'moniteur_federal2']
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                elif section.nom == 'encadrant':
                    # Encadrant : initiateurs et moniteurs
                    eleves = Adherent.objects.filter(
                        statut='eleve',
                        niveau__in=['initiateur1', 'initiateur2', 'moniteur_federal1', 'moniteur_federal2']
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
                else:
                    # Par défaut : tous les élèves
                    eleves = Adherent.objects.filter(
                        statut='eleve'
                    ).values('id', 'nom', 'prenom').order_by('nom', 'prenom')
            
            return JsonResponse({'eleves': list(eleves)})
        except Section.DoesNotExist:
            return JsonResponse({'eleves': []})
    return JsonResponse({'eleves': []})

@require_GET
def api_membres_app(request):
    q = request.GET.get('q', '').strip()
    membres = Adherent.objects.filter(type_personne='adherent')
    if q:
        membres = membres.filter(Q(nom__icontains=q) | Q(prenom__icontains=q))
    membres = membres.order_by('nom', 'prenom')
    data = [
        {
            'id': m.id,
            'nom': m.nom,
            'prenom': m.prenom,
            'niveau': m.get_niveau_display(),
            'sections': [s.get_nom_display() for s in m.sections.all()],
            'statut': m.get_statut_display(),
            'date_delivrance_caci': m.date_delivrance_caci.strftime('%d/%m/%Y') if m.date_delivrance_caci else '',
        }
        for m in membres
    ]
    return JsonResponse({'membres': data})

@require_GET
def api_recherche_non_membre(request):
    nom = request.GET.get('nom', '').strip()
    prenom = request.GET.get('prenom', '').strip()
    non_membres = Adherent.objects.filter(type_personne='non_adherent')
    if nom:
        non_membres = non_membres.filter(nom__icontains=nom)
    if prenom:
        non_membres = non_membres.filter(prenom__icontains=prenom)
    non_membres = non_membres.order_by('nom', 'prenom')
    results = []
    for nm in non_membres:
        results.append({
            'id': nm.id,
            'nom': nm.nom,
            'prenom': nm.prenom,
            'email': nm.email,
            'telephone': nm.telephone,
            'niveau': nm.niveau,
            'statut': nm.statut,
            'sections': [s.id for s in nm.sections.all()],
            'date_naissance': nm.date_naissance.strftime('%Y-%m-%d') if nm.date_naissance else '',
            'adresse': nm.adresse,
            'code_postal': nm.code_postal,
            'ville': nm.ville,
            'numero_licence': nm.numero_licence,
            'assurance': nm.assurance,
            'date_delivrance_caci': nm.date_delivrance_caci.strftime('%Y-%m-%d') if nm.date_delivrance_caci else '',
            'caci_fichier_url': nm.caci_fichier.url if nm.caci_fichier else '',
        })
    return JsonResponse({'results': results})

# Vues pour l'import Excel
@login_required
def import_adherents_excel(request):
    """Import en masse d'adhérents depuis un fichier Excel"""
    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            
            # Vérifier l'extension du fichier
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Veuillez sélectionner un fichier Excel (.xlsx ou .xls)')
                return render(request, 'gestion/import_adherents_excel.html')
            
            try:
                import pandas as pd
                from datetime import datetime
                import re
                
                # Lire le fichier Excel
                if excel_file.name.endswith('.xlsx'):
                    df = pd.read_excel(excel_file, engine='openpyxl')
                else:
                    df = pd.read_excel(excel_file, engine='xlrd')
                
                # --- Ajout des nouveaux champs dans l'import ---
                required_columns = ['nom', 'prenom', 'date_naissance', 'adresse', 'code_postal', 'ville', 'email', 
                                  'telephone', 'numero_licence', 'assurance', 'date_delivrance_caci', 'niveau', 'statut']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f'Colonnes manquantes dans le fichier : {", ".join(missing_columns)}')
                    return render(request, 'gestion/import_adherents_excel.html')
                
                # Valeurs autorisées pour la validation
                niveaux_valides = [c[0] for c in Adherent.NIVEAUX_CHOICES]
                statuts_valides = [c[0] for c in Adherent.STATUT_CHOICES]
                assurances_valides = [c[0] for c in Adherent.ASSURANCE_CHOICES]
                
                # Traitement des données
                success_count = 0
                error_count = 0
                errors = []
                imported_adherents = []
                
                for index, row in df.iterrows():
                    try:
                        # Nettoyer les données
                        nom = str(row['nom']).strip() if pd.notna(row['nom']) else ''
                        prenom = str(row['prenom']).strip() if pd.notna(row['prenom']) else ''
                        email = str(row['email']).strip() if pd.notna(row['email']) else ''
                        numero_licence = str(row['numero_licence']).strip() if pd.notna(row['numero_licence']) else ''
                        assurance = str(row['assurance']).strip() if pd.notna(row['assurance']) else ''
                        code_postal = str(row['code_postal']).strip() if pd.notna(row['code_postal']) else ''
                        ville = str(row['ville']).strip() if pd.notna(row['ville']) else ''
                        
                        # Vérifier les champs obligatoires
                        if not nom or not prenom or not email:
                            errors.append(f"Ligne {index + 2}: nom, prénom et email sont obligatoires")
                            error_count += 1
                            continue
                        
                        # Vérifier si l'adhérent existe déjà (email OU nom+prénom)
                        if Adherent.objects.filter(email=email).exists():
                            errors.append(f"Ligne {index + 2}: l'email {email} existe déjà")
                            error_count += 1
                            continue
                        
                        if Adherent.objects.filter(nom=nom, prenom=prenom).exists():
                            errors.append(f"Ligne {index + 2}: l'adhérent {nom} {prenom} existe déjà")
                            error_count += 1
                            continue
                        
                        # Validation du format des dates (DD/MM/YYYY uniquement)
                        date_naissance = None
                        date_delivrance_caci = None
                        
                        # Validation date de naissance
                        if pd.notna(row['date_naissance']):
                            date_str = str(row['date_naissance']).strip()
                            if re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
                                try:
                                    date_naissance = datetime.strptime(date_str, '%d/%m/%Y').date()
                                except ValueError:
                                    errors.append(f"Ligne {index + 2}: date de naissance '{date_str}' invalide (format DD/MM/YYYY attendu)")
                                    error_count += 1
                                    continue
                            else:
                                errors.append(f"Ligne {index + 2}: date de naissance '{date_str}' format incorrect (DD/MM/YYYY attendu)")
                                error_count += 1
                                continue
                        else:
                            date_naissance = datetime.now().date()
                        
                        # Validation date fin validité CACI
                        if pd.notna(row['date_delivrance_caci']):
                            date_str = str(row['date_delivrance_caci']).strip()
                            if re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
                                try:
                                    date_delivrance_caci = datetime.strptime(date_str, '%d/%m/%Y').date()
                                except ValueError:
                                    errors.append(f"Ligne {index + 2}: date fin validité CACI '{date_str}' invalide (format DD/MM/YYYY attendu)")
                                    error_count += 1
                                    continue
                            else:
                                errors.append(f"Ligne {index + 2}: date fin validité CACI '{date_str}' format incorrect (DD/MM/YYYY attendu)")
                                error_count += 1
                                continue
                        else:
                            date_delivrance_caci = datetime.now().date()
                        
                        # Validation du niveau
                        niveau = str(row['niveau']).strip() if pd.notna(row['niveau']) else 'debutant'
                        if niveau not in niveaux_valides:
                            errors.append(f"Ligne {index + 2}: niveau '{niveau}' invalide (valeurs: {', '.join(niveaux_valides)})")
                            error_count += 1
                            continue
                        
                        # Validation du statut
                        statut = str(row['statut']).strip() if pd.notna(row['statut']) else 'eleve'
                        if statut not in statuts_valides:
                            errors.append(f"Ligne {index + 2}: statut '{statut}' invalide (valeurs: {', '.join(statuts_valides)})")
                            error_count += 1
                            continue
                        
                        # Validation de l'assurance
                        if assurance not in assurances_valides:
                            errors.append(f"Ligne {index + 2}: assurance '{assurance}' invalide (valeurs: {', '.join(assurances_valides)})")
                            error_count += 1
                            continue
                        
                        # Validation des sections
                        sections_invalides = []
                        if 'sections' in df.columns and pd.notna(row['sections']):
                            sections_str = str(row['sections']).strip()
                            if sections_str:
                                section_names = [s.strip() for s in sections_str.split(',')]
                                for section_name in section_names:
                                    if not Section.objects.filter(nom=section_name).exists():
                                        sections_invalides.append(section_name)
                        
                        if sections_invalides:
                            errors.append(f"Ligne {index + 2}: sections inexistantes: {', '.join(sections_invalides)}")
                            error_count += 1
                            continue
                        
                        # Si toutes les validations passent, créer l'adhérent
                        adherent = Adherent.objects.create(
                            nom=nom,
                            prenom=prenom,
                            date_naissance=date_naissance,
                            adresse=str(row['adresse']).strip() if pd.notna(row['adresse']) else '',
                            code_postal=code_postal,
                            ville=ville,
                            email=email,
                            telephone=str(row['telephone']).strip() if pd.notna(row['telephone']) else '',
                            numero_licence=numero_licence,
                            assurance=assurance,
                            date_delivrance_caci=date_delivrance_caci,
                            niveau=niveau,
                            statut=statut,
                        )
                        
                        # Ajouter les sections
                        if 'sections' in df.columns and pd.notna(row['sections']):
                            sections_str = str(row['sections']).strip()
                            if sections_str:
                                section_names = [s.strip() for s in sections_str.split(',')]
                                for section_name in section_names:
                                    try:
                                        section = Section.objects.get(nom=section_name)
                                        adherent.sections.add(section)
                                    except Section.DoesNotExist:
                                        pass
                        
                        success_count += 1
                        imported_adherents.append(f"{nom} {prenom}")
                        
                    except Exception as e:
                        errors.append(f"Ligne {index + 2}: {str(e)}")
                        error_count += 1
                
                # Stocker les résultats dans la session pour l'affichage
                request.session['import_results'] = {
                    'success_count': success_count,
                    'error_count': error_count,
                    'errors': errors,
                    'imported_adherents': imported_adherents
                }
                
                # Messages de résultat
                if success_count > 0:
                    messages.success(request, f'{success_count} adhérent(s) importé(s) avec succès')
                
                if error_count > 0:
                    messages.warning(request, f'{error_count} erreur(s) lors de l\'import')
                
                return render(request, 'gestion/import_adherents_excel.html')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de la lecture du fichier Excel : {str(e)}')
                return render(request, 'gestion/import_adherents_excel.html')
        else:
            messages.error(request, 'Aucun fichier sélectionné')
            return render(request, 'gestion/import_adherents_excel.html')
    
    # Affichage de la page d'import (GET request)
    return render(request, 'gestion/import_adherents_excel.html')

@login_required
def download_excel_template(request):
    """Télécharger le modèle Excel pour l'import d'adhérents"""
    import pandas as pd
    from datetime import date, timedelta
    
    # --- Ajout des champs dans le fichier exemple ---
    data = {
        'nom': ['Dupont', 'Martin', 'Bernard'],
        'prenom': ['Jean', 'Marie', 'Pierre'],
        'date_naissance': ['15/05/1990', '03/12/1985', '22/08/1992'],
        'adresse': ['123 Rue de la Paix, Paris', '456 Avenue des Champs, Lyon', '789 Boulevard de la Mer, Nice'],
        'code_postal': ['75001', '69002', '06000'],
        'ville': ['Paris', 'Lyon', 'Nice'],
        'email': ['jean.dupont@email.com', 'marie.martin@email.com', 'pierre.bernard@email.com'],
        'telephone': ['0123456789', '0987654321', '0567891234'],
        'numero_licence': ['123456', '', '789101'],
        'assurance': ['Piscine', 'Loisir 1', ''],
        'date_delivrance_caci': ['31/12/2025', '30/06/2026', '15/09/2025'],
        'niveau': ['niveau1', 'niveau2', 'debutant'],
        'statut': ['eleve', 'eleve', 'eleve'],
        'sections': ['bapteme,prepa_niveau1', 'prepa_niveau2', 'bapteme']
    }
    
    df = pd.DataFrame(data)
    
    # Créer la réponse HTTP
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modele_import_adherents.xlsx"'
    
    # Écrire le fichier Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Adhérents', index=False)
        
        # Ajouter une feuille avec les instructions
        instructions = pd.DataFrame({
            'Colonne': ['nom', 'prenom', 'date_naissance', 'adresse', 'code_postal', 'ville', 'email', 'telephone', 'numero_licence', 'assurance', 'date_delivrance_caci', 'niveau', 'statut', 'sections'],
            'Description': [
                'Nom de famille (obligatoire)',
                'Prénom (obligatoire)',
                'Date de naissance (format: DD/MM/YYYY)',
                'Adresse complète',
                'Code postal',
                'Ville',
                'Email (obligatoire, unique)',
                'Numéro de téléphone',
                'Numéro de licence (facultatif)',
                "Assurance (valeurs possibles : " + ', '.join([c[0] for c in Adherent.ASSURANCE_CHOICES]) + ")",
                'Date de délivrance du CACI (format: DD/MM/YYYY)',
                "Niveau de plongée (valeurs possibles : " + ', '.join([c[0] for c in Adherent.NIVEAUX_CHOICES]) + ")",
                "Statut (valeurs possibles : " + ', '.join([c[0] for c in Adherent.STATUT_CHOICES]) + ")",
                "Sections (séparées par des virgules : " + ', '.join([c[0] for c in Section.SECTIONS_CHOICES]) + ")"
            ],
            'Obligatoire': ['Oui', 'Oui', 'Non', 'Non', 'Non', 'Non', 'Oui', 'Non', 'Non', 'Non', 'Non', 'Non', 'Non', 'Non'],
            'Exemple': ['Dupont', 'Jean', '15/05/1990', '123 Rue de la Paix, Paris', '75001', 'Paris', 'jean.dupont@email.com', '0123456789', '123456', 'Piscine', '31/12/2025', 'niveau1', 'eleve', 'bapteme,prepa_niveau1']
        })
        
        instructions.to_excel(writer, sheet_name='Instructions', index=False)
    
    return response

# Vues d'authentification personnalisées
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        user = self.request.user
        if hasattr(user, 'adherent_profile') and getattr(user.adherent_profile, 'statut', None) == 'eleve':
            return reverse_lazy('suivi_formation_eleve', kwargs={'eleve_id': user.adherent_profile.pk})
        return super().get_success_url()

class CustomLogoutView(LogoutView):
    next_page = 'login'

# Vues pour les lieux
@method_decorator(group_required('admin'), name='dispatch')
class LieuListView(LoginRequiredMixin, ListView):
    model = Lieu
    template_name = 'gestion/lieu_list.html'
    context_object_name = 'lieux'

@method_decorator(group_required('admin'), name='dispatch')
class LieuCreateView(LoginRequiredMixin, CreateView):
    model = Lieu
    fields = ['nom', 'adresse', 'code_postal', 'ville']
    template_name = 'gestion/lieu_form.html'
    success_url = reverse_lazy('lieu_list')
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.order_fields(['nom', 'adresse', 'code_postal', 'ville'])
        return form

@method_decorator(group_required('admin'), name='dispatch')
class LieuUpdateView(LoginRequiredMixin, UpdateView):
    model = Lieu
    fields = ['nom', 'adresse', 'code_postal', 'ville']
    template_name = 'gestion/lieu_form.html'
    success_url = reverse_lazy('lieu_list')
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.order_fields(['nom', 'adresse', 'code_postal', 'ville'])
        return form

@method_decorator(group_required('admin'), name='dispatch')
class LieuDeleteView(LoginRequiredMixin, DeleteView):
    model = Lieu
    template_name = 'gestion/lieu_confirm_delete.html'
    success_url = reverse_lazy('lieu_list')

@login_required
def generer_lien_inscription_seance(request, seance_id):
    seance = get_object_or_404(Seance, pk=seance_id)
    # Calcul du mercredi précédent la date de la séance
    seance_date = seance.date
    weekday = seance_date.weekday()  # 0=lundi, 2=mercredi
    days_since_wed = (weekday - 2) % 7
    expiration_date = seance_date - timedelta(days=days_since_wed)
    # Heure d'expiration : 23:59
    expiration = timezone.make_aware(datetime.combine(expiration_date, datetime.max.time().replace(hour=23, minute=59, second=0, microsecond=0)))
    lien, created = LienInscriptionSeance.objects.update_or_create(
        seance=seance,
        defaults={'date_expiration': expiration}
    )
    messages.success(request, "Lien d'inscription généré !")
    return redirect('seance_detail', pk=seance.pk)

@login_required
def envoyer_mail_invitation_seance(request, seance_id):
    seance = get_object_or_404(Seance, pk=seance_id)
    lien = seance.liens_inscription.last()
    if not lien:
        messages.error(request, "Aucun lien d'inscription généré pour cette séance.")
        return redirect('seance_detail', pk=seance_id)
    # Récupère tous les adhérents du club
    adherents = Adherent.objects.filter(type_personne='adherent')
    emails = [a.email for a in adherents if a.email]
    if not emails:
        messages.error(request, "Aucun email d'adhérent trouvé.")
        return redirect('seance_detail', pk=seance_id)
    # Prépare le message
    subject = f"Inscription à la séance du {seance.date.strftime('%d/%m/%Y')} - {seance.lieu.nom}"
    url = request.build_absolute_uri(f"/inscription/{lien.uuid}/")
    message_txt = render_to_string('gestion/email_invitation_seance.txt', {'seance': seance, 'lien': lien, 'url': url})
    message_html = render_to_string('gestion/email_invitation_seance.html', {'seance': seance, 'lien': lien, 'url': url})
    cc_emails = getattr(settings, 'EMAIL_CC_DEFAULT', [])
    datatuple = [(subject, adherent, [adherent.email], cc_emails) for adherent in adherents if adherent.email]
    from django.core.mail import EmailMultiAlternatives
    for subject, adherent, recipient_list, cc_list in datatuple:
        url = request.build_absolute_uri(f"/inscription/{lien.uuid}/")
        context = {'seance': seance, 'lien': lien, 'url': url, 'adherent': adherent}
        message_txt = render_to_string('gestion/email_invitation_seance.txt', context)
        message_html = render_to_string('gestion/email_invitation_seance.html', context)
        email = EmailMultiAlternatives(subject, message_txt, None, recipient_list, cc=cc_list)
        email.attach_alternative(message_html, "text/html")
        # Attacher l'image de signature en inline
        signature_img_path = os.path.join(settings.BASE_DIR, 'static', 'Signature_mouss2.png')
        if os.path.exists(signature_img_path):
            with open(signature_img_path, 'rb') as img:
                mime_img = MIMEImage(img.read(), _subtype='png')
                mime_img.add_header('Content-ID', '<signature_mouss2>')
                mime_img.add_header('Content-Disposition', 'inline', filename='Signature_mouss2.png')
                email.attach(mime_img)
        email.send()
    messages.success(request, f"Invitation envoyée à {len(emails)} adhérents.")
    return redirect('seance_detail', pk=seance_id)


def inscription_seance_uuid(request, uuid):
    lien = get_object_or_404(LienInscriptionSeance, uuid=uuid)
    now = timezone.now()
    if now > lien.date_expiration:
        return render(request, 'gestion/inscription_expiree.html', {'seance': lien.seance})
    return render(request, 'gestion/inscription_seance.html', {'seance': lien.seance, 'lien': lien})

@csrf_exempt
@require_POST
def api_inscrire_membre_app(request):
    import json
    try:
        data = json.loads(request.body)
        uuid = data.get('uuid')
        membre_id = data.get('membre_id')
        lien = LienInscriptionSeance.objects.get(uuid=uuid)
        seance = lien.seance
        membre = Adherent.objects.get(id=membre_id, type_personne='adherent')
        from .models import InscriptionSeance
        inscription = InscriptionSeance.objects.filter(seance=seance, personne=membre).first()
        if not inscription:
            InscriptionSeance.objects.create(
                seance=seance,
                personne=membre,
                covoiturage=data.get('covoiturage', ''),
                lieu_covoiturage=data.get('lieu_covoiturage', '')
            )
        else:
            inscription.covoiturage = data.get('covoiturage', '')
            inscription.lieu_covoiturage = data.get('lieu_covoiturage', '')
            inscription.save()
        # Envoi mail confirmation
        if membre.email:
            subject = f"Confirmation d'inscription à la séance du {seance.date.strftime('%d/%m/%Y')}"
            url = request.build_absolute_uri(f"/inscription/{lien.uuid}/")
            message = render_to_string('gestion/email_confirmation_inscription.txt', {'seance': seance, 'personne': membre, 'url': url})
            cc_emails = getattr(settings, 'EMAIL_CC_DEFAULT', [])
            from django.core.mail import EmailMessage
            email = EmailMessage(subject, message, None, [membre.email])#
            email.send(fail_silently=True)
        return JsonResponse({'success': True, 'message': 'Inscription réussie ! Un email de confirmation vous a été envoyé.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
@require_POST
def api_inscrire_non_membre(request):
    uuid = request.POST.get('uuid')
    lien = get_object_or_404(LienInscriptionSeance, uuid=uuid)
    seance = lien.seance
    nom = request.POST.get('nom', '').strip()
    prenom = request.POST.get('prenom', '').strip()
    non_membre = Adherent.objects.filter(type_personne='non_adherent', nom__iexact=nom, prenom__iexact=prenom).first()
    from .models import InscriptionSeance
    from .forms import PublicNonAdherentInscriptionForm
    debug_msgs = []
    debug_msgs.append(f"POST: {dict(request.POST)}")
    if non_membre and InscriptionSeance.objects.filter(seance=seance, personne=non_membre).exists():
        debug_msgs.append("Déjà inscrit")
        return JsonResponse({'success': False, 'message': 'Vous êtes déjà inscrit à cette séance.', 'debug': debug_msgs})
    if non_membre:
        form = PublicNonAdherentInscriptionForm(request.POST, request.FILES, instance=non_membre)
        debug_msgs.append("Form instance existant (AdherentPublicForm)")
    else:
        form = PublicNonAdherentInscriptionForm(request.POST, request.FILES)
        debug_msgs.append("Form nouvelle instance (AdherentPublicForm)")
    if form.is_valid():
        personne = form.save(commit=False)
        personne.type_personne = 'non_adherent'
        personne.save()
        form.save_m2m()
        debug_msgs.append(f"Formulaire valide, personne id={personne.id}")
        # Inscription à la séance
        covoiturage = form.cleaned_data.get('covoiturage', '')
        lieu_covoiturage = form.cleaned_data.get('lieu_covoiturage', '')
        if not InscriptionSeance.objects.filter(seance=seance, personne=personne).exists():
            InscriptionSeance.objects.create(
                seance=seance,
                personne=personne,
                covoiturage=covoiturage,
                lieu_covoiturage=lieu_covoiturage
            )
            debug_msgs.append("Inscription créée")
        else:
            inscription = InscriptionSeance.objects.filter(seance=seance, personne=personne).first()
            inscription.covoiturage = covoiturage
            inscription.lieu_covoiturage = lieu_covoiturage
            inscription.save()
            debug_msgs.append("Inscription mise à jour")
        # Envoi mail confirmation
        if personne.email:
            subject = f"Confirmation d'inscription à la séance du {seance.date.strftime('%d/%m/%Y')}"
            url = request.build_absolute_uri(f"/inscription/{lien.uuid}/")
            message = render_to_string('gestion/email_confirmation_inscription.txt', {'seance': seance, 'personne': personne, 'url': url})
            cc_emails = getattr(settings, 'EMAIL_CC_DEFAULT', [])
            from django.core.mail import EmailMessage
            try:
                email = EmailMessage(subject, message, None, [personne.email])#, cc=cc_emails
                email.send(fail_silently=True)
                debug_msgs.append("Mail envoyé")
            except Exception as mail_e:
                debug_msgs.append(f"Erreur envoi mail: {str(mail_e)}")
                print(f"[INSCRIPTION] Erreur envoi mail: {str(mail_e)}") # Log to console
        return JsonResponse({'success': True, 'message': 'Inscription réussie ! Un email de confirmation vous a été envoyé.', 'debug': debug_msgs})
    else:
        debug_msgs.append(f"Form errors: {form.errors}")
        return JsonResponse({'success': False, 'message': 'Erreur : ' + str(form.errors), 'debug': debug_msgs})

@login_required
def supprimer_inscription_seance(request, inscription_id):
    from .models import InscriptionSeance
    inscription = get_object_or_404(InscriptionSeance, id=inscription_id)
    seance_id = inscription.seance.id
    if request.method == 'POST':
        inscription.delete()
        messages.success(request, "Inscription supprimée avec succès.")
        return redirect('seance_detail', pk=seance_id)
    return render(request, 'gestion/inscription_confirm_delete.html', {'inscription': inscription})

@login_required
def exporter_inscrits_seance(request, seance_id):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    seance = get_object_or_404(Seance, pk=seance_id)
    inscrits = seance.inscriptions.select_related('personne').all()
    # Sépare encadrants et autres
    encadrants = [i.personne for i in inscrits if i.personne.statut == 'encadrant']
    eleves = [i.personne for i in inscrits if i.personne.statut != 'encadrant']
    # Création du workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inscrits séance'
    # Ligne 1 : Date
    ws.append([f"Date : {seance.date.strftime('%d/%m/%Y')}"])
    # Ligne 2 : Encadrants
    row2 = [None, None, None, 'Encadrant']
    for enc in encadrants:
        row2.append(f"{enc.nom} {enc.prenom}")
    ws.append(row2)
    # Ligne 3 : En-têtes
    headers = ['Nom Prénom', 'Niveau', 'Section', 'Profondeur max']
    headers += ['' for _ in encadrants]
    ws.append(headers)
    # Lignes suivantes : inscrits
    for eleve in eleves:
        nom_prenom = f"{eleve.nom} {eleve.prenom}"
        niveau = eleve.get_niveau_display() if hasattr(eleve, 'get_niveau_display') else ''
        sections = ', '.join([s.get_nom_display() for s in eleve.sections.all()])
        row = [nom_prenom, niveau, sections, '']
        row += ['' for _ in encadrants]
        ws.append(row)
    # Mise en forme : texte vertical pour les encadrants
    for idx, enc in enumerate(encadrants, start=5):
        cell = ws.cell(row=2, column=idx)
        cell.alignment = Alignment(text_rotation=90, vertical='bottom', horizontal='center')
    # Largeur des colonnes
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    # Export
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="inscrits_seance_{seance.id}.xlsx"'
    wb.save(response)
    return response

@login_required
def exporter_covoiturage_seance(request, seance_id):
    import pandas as pd
    seance = get_object_or_404(Seance, pk=seance_id)
    inscrits = seance.inscriptions.select_related('personne').all()
    data = []
    for ins in inscrits:
        data.append({
            'Nom': ins.personne.nom,
            'Prénom': ins.personne.prenom,
            'Type': 'Adhérent' if ins.personne.type_personne == 'adherent' else 'Non adhérent',
            'Covoiturage': dict(ins.COVOITURAGE_CHOICES).get(ins.covoiturage, ''),
            'Lieu de prise en charge': ins.lieu_covoiturage or '',
        })
    df = pd.DataFrame(data)
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="covoiturage_seance_{seance.id}.xlsx"'
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return response

@method_decorator(csrf_protect, name='dispatch')
class AdherentPublicCreateView(CreateView):
    model = Adherent
    form_class = AdherentPublicForm
    template_name = 'gestion/adherent_public_form.html'
    success_url = None  # Pas de redirection

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inscription_2025_2026'] = getattr(self, 'inscription_2025_2026', False)
        return context

    def dispatch(self, request, *args, **kwargs):
        self.inscription_2025_2026 = kwargs.get('inscription_2025_2026', False)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.save()
        return self.render_to_response(self.get_context_data(form=self.form_class(), inscription_success=True))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form, inscription_success=False))

@method_decorator(group_required('admin'), name='dispatch')
class ExerciceListView(LoginRequiredMixin, ListView):
    model = Exercice
    template_name = 'gestion/exercice_list.html'
    context_object_name = 'exercices'

@method_decorator(group_required('admin'), name='dispatch')
class ExerciceCreateView(LoginRequiredMixin, CreateView):
    model = Exercice
    form_class = ExerciceForm
    template_name = 'gestion/exercice_form.html'
    success_url = reverse_lazy('exercice_list')

@method_decorator(group_required('admin'), name='dispatch')
class ExerciceUpdateView(LoginRequiredMixin, UpdateView):
    model = Exercice
    form_class = ExerciceForm
    template_name = 'gestion/exercice_form.html'
    success_url = reverse_lazy('exercice_list')

@method_decorator(group_required('admin'), name='dispatch')
class ExerciceDeleteView(LoginRequiredMixin, DeleteView):
    model = Exercice
    template_name = 'gestion/exercice_confirm_delete.html'
    success_url = reverse_lazy('exercice_list')

@login_required
def export_adherents_excel(request):
    import pandas as pd
    adherents = Adherent.objects.all().order_by('nom', 'prenom')
    data = []
    for a in adherents:
        data.append({
            'Nom': a.nom.upper(),
            'Prénom': a.prenom.capitalize(),
            'Email': a.email,
            'Téléphone': a.telephone,
            'Adresse': a.adresse,
            'Code postal': a.code_postal,
            'Ville': a.ville,
            'Numéro de licence': a.numero_licence,
            'Assurance': a.assurance,
            'Date délivrance CACI': a.date_delivrance_caci,
            'Niveau': a.get_niveau_display(),
            'Statut': a.get_statut_display(),
        })
    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="adherents.xlsx"'
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return response

@login_required
def importer_palanquees_seance(request, seance_id):
    import openpyxl
    seance = get_object_or_404(Seance, pk=seance_id)
    if request.method == 'POST' and request.FILES.get('fichier_palanquees'):
        fichier = request.FILES['fichier_palanquees']
        try:
            wb = openpyxl.load_workbook(fichier)
            ws = wb.active
            # Suppression des palanquées existantes
            seance.palanques.all().delete()
            # Lecture des encadrants (ligne 2, colonnes >= 5)
            encadrants = []
            for col in range(5, ws.max_column+1):
                val = ws.cell(row=2, column=col).value
                if val:
                    nom, *prenom = val.split()
                    prenom = ' '.join(prenom)
                    from gestion.models import Adherent
                    enc = Adherent.objects.filter(nom__iexact=nom.strip(), prenom__iexact=prenom.strip(), statut='encadrant').first()
                    encadrants.append(enc)
                else:
                    encadrants.append(None)
            # Lecture des élèves (lignes >= 4)
            for col_idx, enc in enumerate(encadrants, start=5):
                if not enc:
                    continue
                eleves = []
                profondeurs = []
                for row in range(4, ws.max_row+1):
                    x = ws.cell(row=row, column=col_idx).value
                    if x and str(x).strip().lower() == 'x':
                        nom_prenom = ws.cell(row=row, column=1).value or ''
                        nom, *prenom = nom_prenom.split()
                        prenom = ' '.join(prenom)
                        eleve = Adherent.objects.filter(nom__iexact=nom.strip(), prenom__iexact=prenom.strip()).first()
                        if eleve:
                            eleves.append(eleve)
                            # Profondeur max (colonne 4)
                            prof = ws.cell(row=row, column=4).value
                            try:
                                profondeurs.append(int(prof))
                            except:
                                pass
                if eleves:
                    from gestion.models import Section, Palanquee
                    # Trouver l'élève avec le niveau le plus faible
                    eleve_min = None
                    niveau_min = None
                    for e in eleves:
                        if hasattr(e, 'niveau') and e.niveau:
                            if niveau_min is None or e.niveau < niveau_min:
                                niveau_min = e.niveau
                                eleve_min = e
                    section = eleve_min.sections.first() if eleve_min else None
                    if not section:
                        messages.error(request, f"Aucune section trouvée pour l'élève de niveau le plus faible dans la palanquée de {enc.nom} {enc.prenom}. Palanquée non créée.")
                        continue
                    profondeur_max = min(profondeurs) if profondeurs else None
                    palanquee = Palanquee.objects.create(
                        nom=f"Palanquée {enc.nom} {enc.prenom}",
                        seance=seance,
                        section=section,
                        encadrant=enc,
                        precision_exercices='',
                        profondeur_max=profondeur_max
                    )
                    palanquee.eleves.set(eleves)
            messages.success(request, "Import des palanquées effectué avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {str(e)}")
    else:
        messages.error(request, "Aucun fichier fourni.")
    return redirect('seance_detail', seance_id)

@login_required
def creer_palanquees(request, seance_id):
    seance = get_object_or_404(Seance, pk=seance_id)
    inscrits = seance.inscriptions.select_related('personne').all()
    eleves = [i.personne for i in inscrits if i.personne.statut == 'eleve']
    encadrants = [i.personne for i in inscrits if i.personne.statut == 'encadrant']

    if request.method == 'POST':
        from django.db import transaction
        affectations = {str(e.id): None for e in eleves}
        palanquees_data = {}
        # Récupérer affectations et champs aptitude
        for eleve in eleves:
            for moniteur in encadrants:
                if request.POST.get(f'affectation_{eleve.id}_{moniteur.id}'):
                    affectations[str(eleve.id)] = moniteur.id
            # aptitude
            palanquees_data[str(eleve.id)] = {
                'aptitude': request.POST.get(f'aptitude_{eleve.id}', '').strip()
            }
        # Récupérer profondeur/durée max par moniteur
        profs = {}
        durees = {}
        for moniteur in encadrants:
            prof = request.POST.get(f'profondeur_max_{moniteur.id}')
            duree = request.POST.get(f'duree_max_{moniteur.id}')
            profs[moniteur.id] = int(prof) if prof else None
            durees[moniteur.id] = int(duree) if duree else None
        # --- Gestion des palanquées autonomes ---
        # Cherche les colonnes autonomes dynamiques
        autonome_cols = []
        for key in request.POST.keys():
            if key.startswith('affectation_autonome_'):
                parts = key.split('_')
                if len(parts) == 4:
                    autonome_num = parts[2]
                    if autonome_num not in autonome_cols:
                        autonome_cols.append(autonome_num)
        # Regroupe les élèves par colonne autonome
        autonomes_groupes = {num: [] for num in autonome_cols}
        for num in autonome_cols:
            for eleve in eleves:
                if request.POST.get(f'affectation_autonome_{num}_{eleve.id}'):
                    autonomes_groupes[num].append(str(eleve.id))
        # --- Fin gestion autonomes ---
        # Vérification : tous les moniteurs doivent avoir au moins un élève
        moniteurs_utilises = set([mid for mid in affectations.values() if mid])
        moniteurs_sans = [m for m in encadrants if m.id not in moniteurs_utilises]
        # Suppression du blocage : on autorise la création même si certains moniteurs n'ont pas d'élève et qu'il n'y a pas de colonne autonome
        # if moniteurs_sans and not autonome_cols:
        #     from django.contrib import messages
        #     messages.error(request, "Tous les moniteurs doivent avoir au moins une palanquée (un élève affecté), ou il doit y avoir au moins une palanquée d'autonomes.")
        #     return render(request, 'gestion/creer_palanquees.html', {
        #         'seance': seance, 'eleves': eleves, 'encadrants': encadrants
        #     })
        # Vérification : élèves non affectés
        eleves_non_aff = [e for e, m in affectations.items() if not m]
        # Retirer les élèves affectés à une palanquée autonome
        for num, eids in autonomes_groupes.items():
            for eid in eids:
                if eid in eleves_non_aff:
                    eleves_non_aff.remove(eid)
        # Suppression de l'avertissement et de la confirmation : on ne bloque plus la création
        # if eleves_non_aff and not request.POST.get('confirm_non_affectes'):
        #     from django.contrib import messages
        #     messages.warning(request, "Certains élèves ne sont pas affectés à une palanquée. Confirmez pour continuer.")
        #     return render(request, 'gestion/creer_palanquees.html', {
        #         'seance': seance, 'eleves': eleves, 'encadrants': encadrants,
        #         'eleves_non_aff': eleves_non_aff, 'demande_confirmation': True
        #     })
        # Création des palanquées (on ajoute, on ne supprime pas les existantes)
        try:
            with transaction.atomic():
                # Regrouper les élèves par moniteur
                groupes = {}
                for eid, mid in affectations.items():
                    if mid:
                        groupes.setdefault(mid, []).append(eid)
                for moniteur in encadrants:
                    if moniteur.id not in groupes:
                        continue
                    eleves_ids = groupes[moniteur.id]
                    eleves_objs = [e for e in eleves if str(e.id) in eleves_ids]
                    # Section la plus faible des élèves
                    section = None
                    niveau_min = None
                    for e in eleves_objs:
                        s = e.sections.first()
                        if s:
                            if not section or (hasattr(e, 'niveau') and (niveau_min is None or e.niveau < niveau_min)):
                                section = s
                                niveau_min = e.niveau
                    palanquee = Palanquee.objects.create(
                        nom=f"Palanquée {moniteur.nom} {moniteur.prenom}",
                        seance=seance,
                        section=section,
                        encadrant=moniteur,
                        precision_exercices='',
                        profondeur_max=profs[moniteur.id],
                        duree=durees[moniteur.id]
                    )
                    for eid in eleves_ids:
                        aptitude = palanquees_data[eid]['aptitude']
                        PalanqueeEleve.objects.create(palanquee=palanquee, eleve_id=eid, aptitude=aptitude)
                # Création des palanquées autonomes
                for num, eids in autonomes_groupes.items():
                    if not eids:
                        continue
                    eleves_objs = [e for e in eleves if str(e.id) in eids]
                    # Section la plus faible des élèves
                    section = None
                    niveau_min = None
                    for e in eleves_objs:
                        s = e.sections.first()
                        if s:
                            if not section or (hasattr(e, 'niveau') and (niveau_min is None or e.niveau < niveau_min)):
                                section = s
                                niveau_min = e.niveau
                    # Récupérer profondeur et durée pour cette colonne autonome
                    prof = request.POST.get(f'profondeur_max_autonome_{num}')
                    duree = request.POST.get(f'duree_max_autonome_{num}')
                    profondeur_max = int(prof) if prof else None
                    duree_max = int(duree) if duree else None
                    palanquee = Palanquee.objects.create(
                        nom=f"Palanquée autonomes n°{num}",
                        seance=seance,
                        section=section,
                        encadrant=None,
                        precision_exercices='',
                        profondeur_max=profondeur_max,
                        duree=duree_max
                    )
                    for eid in eids:
                        aptitude = palanquees_data[eid]['aptitude']
                        PalanqueeEleve.objects.create(palanquee=palanquee, eleve_id=eid, aptitude=aptitude)
            from django.contrib import messages
            messages.success(request, "Palanquées créées avec succès.")
            return redirect('seance_detail', seance_id)
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"Erreur lors de la création des palanquées : {e}")
            return render(request, 'gestion/creer_palanquees.html', {
                'seance': seance, 'eleves': eleves, 'encadrants': encadrants
            })
    else:
        eleves = sorted(eleves, key=lambda e: (e.nom.upper(), e.prenom.upper()))
        encadrants = sorted(encadrants, key=lambda e: (e.nom.upper(), e.prenom.upper()))
        return render(request, 'gestion/creer_palanquees.html', {
            'seance': seance,
            'eleves': eleves,
            'encadrants': encadrants,
        })

class PalanqueeDeleteView(LoginRequiredMixin, DeleteView):
    model = Palanquee
    template_name = 'gestion/palanquee_confirm_delete.html'
    def get_success_url(self):
        return reverse_lazy('seance_detail', kwargs={'pk': self.object.seance.pk})

@login_required
def generer_fiche_securite(request, seance_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from io import BytesIO
    seance = get_object_or_404(Seance, pk=seance_id)
    palanquees = seance.palanques.select_related('encadrant').prefetch_related('eleves')
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 2*cm
    p.setFont("Helvetica-Bold", 16)
    p.drawString(2*cm, y, f"FICHE DE SÉCURITÉ CLUB")
    y -= 1*cm
    p.setFont("Helvetica", 12)
    p.drawString(2*cm, y, f"Date : {seance.date.strftime('%d/%m/%Y')}  Heure début : {seance.heure_debut.strftime('%Hh%M') if seance.heure_debut else '-'}  Heure fin : {seance.heure_fin.strftime('%Hh%M') if seance.heure_fin else '-'}")
    y -= 1*cm
    p.drawString(2*cm, y, f"Lieu : {seance.lieu}")
    y -= 1*cm
    if seance.directeur_plongee:
        p.drawString(2*cm, y, f"Directeur de plongée : {seance.directeur_plongee.nom_complet}")
        y -= 1*cm
    p.setFont("Helvetica-Bold", 14)
    p.drawString(2*cm, y, "Palanquées :")
    y -= 0.7*cm
    p.setFont("Helvetica", 12)
    for palanquee in palanquees:
        if y < 5*cm:
            p.showPage()
            y = height - 2*cm
        p.setFont("Helvetica-Bold", 12)
        p.drawString(2*cm, y, f"Palanquée : {palanquee.nom}")
        y -= 0.5*cm
        p.setFont("Helvetica", 12)
        p.drawString(2.5*cm, y, f"Encadrant : {palanquee.encadrant.nom_complet} (Niveau : {palanquee.encadrant.get_niveau_display() if hasattr(palanquee.encadrant, 'get_niveau_display') else ''})")
        y -= 0.5*cm
        eleves = ", ".join([f"{e.nom} {e.prenom} ({e.get_niveau_display() if hasattr(e, 'get_niveau_display') else ''})" for e in palanquee.eleves.all()])
        p.drawString(2.5*cm, y, f"Élèves : {eleves if eleves else '-'}")
        y -= 0.5*cm
        profondeur = palanquee.profondeur_max if palanquee.profondeur_max else '-'
        duree = palanquee.duree if palanquee.duree else '-'
        p.drawString(2.5*cm, y, f"Profondeur max : {profondeur} m   Durée : {duree} min")
        y -= 0.7*cm
        p.line(2*cm, y, width-2*cm, y)
        y -= 0.5*cm
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
def generer_fiche_securite_excel(request, seance_id):
    import openpyxl
    from openpyxl.styles import PatternFill
    from io import BytesIO
    seance = get_object_or_404(Seance, pk=seance_id)
    palanquees = list(seance.palanques.select_related('encadrant').prefetch_related('eleves'))
    wb = openpyxl.load_workbook('fiche_securite_modele.xlsx')
    ws = wb.active

    # --- Champs fixes ---
    ws['D2'] = seance.date.strftime('%d/%m/%Y')
    ws['P2'] = seance.heure_debut.strftime('%Hh%M') if seance.heure_debut else "-"
    ws['AA2'] = seance.heure_fin.strftime('%Hh%M') if seance.heure_fin else "-"
    directeur = seance.directeur_plongee.nom_complet if seance.directeur_plongee else "-"
    ws['H5'] = directeur
    ws['H56'] = directeur
    # Présence du président
    ws['AC4'] = "Oui" if getattr(seance, 'presence_president', False) else "Non"

    # --- Mapping blocs palanquée ---
    bloc_map = [
        (18, 'A'), (18, 'M'), (18, 'Y'),
        (31, 'A'), (31, 'M'), (31, 'Y'),
        (44, 'A'), (44, 'M'), (44, 'Y'),
    ]
    niveau_col = {'A': 'J', 'M': 'V', 'Y': 'AH'}
    prof_col = {'A': 'F', 'M': 'R', 'Y': 'AD'}
    duree_col = {'A': 'J', 'M': 'V', 'Y': 'AH'}

    def niveau_encadrant_display(niveau):
        mapping = {
            'encadrant1': 'E1',
            'encadrant2': 'E2',
            'initiateur1': 'E1',
            'initiateur2': 'E2',
            'moniteur_federal1': 'E3',
            'moniteur_federal2': 'E4',
        }
        return mapping.get(niveau, niveau)

    # Découpage en tranches de 9 palanquées
    nb_blocs = (len(palanquees) + 8) // 9
    for bloc_idx in range(nb_blocs):
        if bloc_idx > 0:
            # Ajoute une nouvelle feuille à partir du modèle
            ws_new = wb.copy_worksheet(ws)
            ws_new.title = f"Fiche {bloc_idx+1}"
            ws = ws_new
            # Vider les cellules des blocs palanquées (9 blocs max)
            for idx in range(9):
                if idx < len(bloc_map):
                    base_row, base_col = bloc_map[idx]
                    # Efface encadrant, niveau, profondeur, durée
                    ws[f'{base_col}{base_row}'] = ""
                    ws[f'{niveau_col[base_col]}{base_row}'] = ""
                    for i in range(4):
                        ws[f'{base_col}{base_row+2+i}'] = ""
                        ws[f'{niveau_col[base_col]}{base_row+2+i}'] = ""
                    ws[f'{prof_col[base_col]}{base_row+7}'] = ""
                    ws[f'{duree_col[base_col]}{base_row+7}'] = ""
        # Remplir les infos fixes à chaque feuille/bloc
        ws['D2'] = seance.date.strftime('%d/%m/%Y')
        ws['P2'] = seance.heure_debut.strftime('%Hh%M') if seance.heure_debut else "-"
        ws['AA2'] = seance.heure_fin.strftime('%Hh%M') if seance.heure_fin else "-"
        ws['H5'] = directeur
        ws['H56'] = directeur
        ws['AC4'] = "Oui" if getattr(seance, 'presence_president', False) else "Non"
        # Palanquées de ce bloc
        palanquees_bloc = palanquees[bloc_idx*9:(bloc_idx+1)*9]
        for idx, palanquee in enumerate(palanquees_bloc):
            base_row, base_col = bloc_map[idx]
            ws[f'{base_col}{base_row}'] = palanquee.encadrant.nom_complet if palanquee.encadrant else "AUTONOMES"
            niveau = palanquee.encadrant.niveau if palanquee.encadrant and hasattr(palanquee.encadrant, 'niveau') else "-"
            ws[f'{niveau_col[base_col]}{base_row}'] = niveau_encadrant_display(niveau)
            eleves = list(palanquee.eleves.all())
            for i in range(4):
                nom_cell = f'{base_col}{base_row+2+i}'
                niv_cell = f'{niveau_col[base_col]}{base_row+2+i}'
                if i < len(eleves):
                    eleve = eleves[i]
                    ws[nom_cell] = f"{eleve.nom} {eleve.prenom}"
                    aptitude = "-"
                    palanquee_eleve = palanquee.palanqueeeleve_set.filter(eleve=eleve).first()
                    if palanquee_eleve and palanquee_eleve.aptitude:
                        aptitude = palanquee_eleve.get_aptitude_display()
                    ws[niv_cell] = aptitude
                else:
                    ws[nom_cell] = ""
                    ws[niv_cell] = ""
            ws[f'{prof_col[base_col]}{base_row+7}'] = palanquee.profondeur_max if palanquee.profondeur_max else "-"
            ws[f'{duree_col[base_col]}{base_row+7}'] = palanquee.duree if palanquee.duree else "-"
        # Comptage adultes/enfants/total (sur la première feuille uniquement)
        if bloc_idx == 0:
            adultes = 0
            enfants = 0
            for palanquee in palanquees:
                # Compter les élèves
                for eleve in palanquee.eleves.all():
                    if hasattr(eleve, 'date_naissance') and eleve.date_naissance:
                        from datetime import date
                        age = (date.today() - eleve.date_naissance).days // 365
                        if age < 18:
                            enfants += 1
                        else:
                            adultes += 1
                    else:
                        adultes += 1
                # Compter l'encadrant s'il existe
                if palanquee.encadrant and hasattr(palanquee.encadrant, 'date_naissance') and palanquee.encadrant.date_naissance:
                    from datetime import date
                    age = (date.today() - palanquee.encadrant.date_naissance).days // 365
                    if age < 18:
                        enfants += 1
                    else:
                        adultes += 1
                elif palanquee.encadrant:
                    adultes += 1
            total = adultes + enfants
            ws['AH57'] = adultes
            ws['AH58'] = enfants
            ws['AH59'] = total

    # Export
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="APP_Fiche-secu_{seance.date.strftime('%Y-%m-%d')}.xlsx"'
    return response

def peut_voir_suivi(user, eleve_id):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    if user.groups.filter(name='admin').exists():
        return True
    if user.groups.filter(name='encadrant').exists():
        return True
    adherent = getattr(user, 'adherent_profile', None)
    return user.groups.filter(name='eleve').exists() and adherent and adherent.id == eleve_id

def suivi_formation_eleve(request, eleve_id):
    if not peut_voir_suivi(request.user, int(eleve_id)):
        # Redirige l'élève vers sa propre fiche, les autres vers la liste des élèves
        if request.user.groups.filter(name='eleve').exists():
            adherent = getattr(request.user, 'adherent_profile', None)
            if adherent:
                return redirect('suivi_formation_eleve', eleve_id=adherent.id)
        elif request.user.groups.filter(name='encadrant').exists():
            return redirect('eleve_list')
        else:
            return redirect('dashboard')
    return _suivi_formation_eleve(request, eleve_id)

def _suivi_formation_eleve(request, eleve_id):
    eleve = get_object_or_404(Adherent, pk=eleve_id)
    # Récupérer toutes les sections de l'élève
    sections = eleve.sections.all()
    # Récupérer tous les groupes de compétences de ses sections
    groupes = GroupeCompetence.objects.filter(section__in=sections).prefetch_related('competences__exercices')
    # Récupérer toutes les évaluations exercices de l'élève
    evals = EvaluationExercice.objects.filter(eleve=eleve)
    evals_dict = {(e.exercice_id): e for e in evals}
    # Historique par exercice
    historiques = {}
    for ex in Exercice.objects.all():
        historiques[ex.id] = list(EvaluationExercice.objects.filter(eleve=eleve, exercice=ex).order_by('-date_evaluation'))
    progression = []
    for groupe in groupes:
        groupe_data = {'groupe': groupe, 'competences': [], 'etoile_groupe': True}
        for comp in groupe.competences.all():
            comp_data = {'competence': comp, 'exercices': [], 'etoile_competence': True}
            for ex in comp.exercices.all():
                eval_ex = evals_dict.get(ex.id)
                etoiles = eval_ex.note if eval_ex else 0
                commentaire = eval_ex.commentaire if eval_ex else ''
                comp_data['exercices'].append({'exercice': ex, 'etoiles': etoiles, 'commentaire': commentaire})
                if etoiles < 3:
                    comp_data['etoile_competence'] = False
            # Si la compétence n'a aucun exercice, elle n'est pas validée
            if not comp_data['exercices']:
                comp_data['etoile_competence'] = False
            groupe_data['competences'].append(comp_data)
            if not comp_data['etoile_competence']:
                groupe_data['etoile_groupe'] = False
        progression.append(groupe_data)
    # Résumé
    nb_groupes = len(progression)
    nb_groupes_valides = sum(1 for g in progression if g['etoile_groupe'])
    nb_competences = sum(len(g['competences']) for g in progression)
    nb_competences_valides = sum(1 for g in progression for c in g['competences'] if c['etoile_competence'])
    context = {
        'eleve': eleve,
        'progression': progression,
        'nb_groupes': nb_groupes,
        'nb_groupes_valides': nb_groupes_valides,
        'nb_competences': nb_competences,
        'nb_competences_valides': nb_competences_valides,
        'evals': evals,  # debug
        'range3': [1, 2, 3],
        'historiques': historiques,
    }
    return render(request, 'gestion/suivi_formation_eleve.html', context)

@login_required
@staff_member_required
def admin_inscription_seance(request, seance_id):
    if not request.user.is_staff:
        return redirect('seance_detail', pk=seance_id)
    seance = get_object_or_404(Seance, pk=seance_id)
    if request.method == 'POST':
        print('--- ADMIN INSCRIPTION DEBUG ---')
        print('POST:', dict(request.POST))
        print('FILES:', dict(request.FILES))
        form = AdminInscriptionSeanceForm(request.POST, request.FILES)
        print('Form is valid:', form.is_valid())
        if form.is_valid():
            adherent = form.cleaned_data.get('adherent')
            print('adherent:', adherent)
            if adherent:
                personne = adherent
            else:
                print('form.cleaned_data:', {k: v for k, v in form.cleaned_data.items()})
                # Créer un non adhérent avec tous les champs du formulaire
                personne, created = Adherent.objects.get_or_create(
                    nom=form.cleaned_data['nom'],
                    prenom=form.cleaned_data['prenom'],
                    email=form.cleaned_data['email'],
                    defaults={
                        'type_personne': 'non_adherent',
                        'date_naissance': form.cleaned_data.get('date_naissance'),
                        'adresse': form.cleaned_data.get('adresse', ''),
                        'code_postal': form.cleaned_data.get('code_postal', ''),
                        'ville': form.cleaned_data.get('ville', ''),
                        'telephone': form.cleaned_data.get('telephone', ''),
                        'numero_licence': form.cleaned_data.get('numero_licence', ''),
                        'assurance': form.cleaned_data.get('assurance', ''),
                        'date_delivrance_caci': form.cleaned_data.get('date_delivrance_caci'),
                        'niveau': form.cleaned_data.get('niveau', ''),
                        'statut': form.cleaned_data.get('statut', 'eleve'),
                    }
                )
                print('Adherent created:', created, 'id:', personne.id)
                # Gérer le fichier CACI et la photo après création/mise à jour
                caci_file = form.cleaned_data.get('caci_fichier')
                print('caci_fichier in cleaned_data:', caci_file)
                if caci_file:
                    personne.caci_fichier = caci_file
                    print('caci_fichier affecté à personne')
                if 'photo' in form.cleaned_data:
                    photo_file = form.cleaned_data.get('photo')
                    print('photo in cleaned_data:', photo_file)
                    if photo_file:
                        personne.photo = photo_file
                        print('photo affectée à personne')
                if not created:
                    for field in ['date_naissance', 'adresse', 'code_postal', 'ville', 'telephone', 'numero_licence', 'assurance', 'date_delivrance_caci', 'niveau', 'statut']:
                        value = form.cleaned_data.get(field)
                        if value:
                            setattr(personne, field, value)
                    personne.type_personne = 'non_adherent'
                try:
                    personne.save()
                    print('personne.save() OK')
                except Exception as e:
                    print('ERREUR personne.save():', e)
            # Créer l'inscription si pas déjà inscrite
            from .models import InscriptionSeance
            covoiturage = form.cleaned_data.get('covoiturage', '')
            lieu_covoiturage = form.cleaned_data.get('lieu_covoiturage', '')
            if not InscriptionSeance.objects.filter(seance=seance, personne=personne).exists():
                InscriptionSeance.objects.create(
                    seance=seance,
                    personne=personne,
                    covoiturage=covoiturage,
                    lieu_covoiturage=lieu_covoiturage
                )
                messages.success(request, f"{personne.nom} {personne.prenom} inscrit à la séance.")
            else:
                messages.info(request, f"{personne.nom} {personne.prenom} est déjà inscrit à cette séance.")
            return redirect('seance_detail', pk=seance_id)
    else:
        form = AdminInscriptionSeanceForm()
    return render(request, 'gestion/admin_inscription_seance.html', {'form': form, 'seance': seance})

@csrf_exempt
@require_POST
@login_required
def dupliquer_exercices_palanquee(request):
    import json
    try:
        data = json.loads(request.body)
        source_id = int(data.get('source'))
        cibles = [int(cid) for cid in data.get('cibles', [])]
        source = Palanquee.objects.get(pk=source_id)
        exercices = list(source.exercices_prevus.all())
        for cible_id in cibles:
            cible = Palanquee.objects.get(pk=cible_id)
            cible.exercices_prevus.set(exercices)
            cible.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def envoyer_pdf_palanquees_encadrants(request, seance_id):
    seance = get_object_or_404(Seance, pk=seance_id)
    palanquees = seance.palanques.all()
    nb_envoyes = 0
    erreurs = []
    cc = getattr(settings, 'EMAIL_CC_DEFAULT', [])
    signature_html = get_signature_html()
    signature_img_path = os.path.join(settings.BASE_DIR, 'static', 'Signature_mouss.png')
    for palanquee in palanquees:
        encadrant = palanquee.encadrant
        if not encadrant or not encadrant.email:
            continue
        # Générer le PDF en mémoire (même contenu que generer_fiche_palanquee_pdf)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=12, spaceBefore=20
        )
        normal_style = styles['Normal']
        encadrant_nom = palanquee.encadrant.nom_complet if palanquee.encadrant else "-"
        elements.append(Paragraph(f"Palanquée : {encadrant_nom}", title_style))
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Informations générales", heading_style))
        elements.append(Paragraph(f"<b>Date :</b> {palanquee.seance.date.strftime('%d/%m/%Y')}", normal_style))
        elements.append(Paragraph(f"<b>Lieu :</b> {palanquee.seance.lieu}", normal_style))
        elements.append(Paragraph(f"<b>Encadrant :</b> {encadrant_nom}", normal_style))
        elements.append(Paragraph(f"<b>Section :</b> {palanquee.section.get_nom_display()}", normal_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Élèves", heading_style))
        eleves_list = [eleve.nom_complet for eleve in palanquee.eleves.all()]
        elements.append(Paragraph(f"<b>Participants :</b> {', '.join(eleves_list)}", normal_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Exercices prévus", heading_style))
        exercices_list = [ex.nom for ex in palanquee.exercices_prevus.all()]
        for i, exercice in enumerate(exercices_list, 1):
            elements.append(Paragraph(f"{i}. {exercice}", normal_style))
        elements.append(Spacer(1, 12))
        if palanquee.precision_exercices:
            elements.append(Paragraph("Nota", heading_style))
            elements.append(Paragraph(palanquee.precision_exercices, normal_style))
        doc.build(elements)
        buffer.seek(0)
        # Préparer et envoyer le mail
        subject = f"Fiche palanquée - {palanquee.nom} ({seance.date})"
        body = render_to_string('gestion/email_pdf_palanquee.txt', {'palanquee': palanquee, 'seance': seance})
        body_html = f"<p>{body.replace(chr(10), '<br>')}</p>{signature_html}"
        email = EmailMultiAlternatives(subject, body, to=[encadrant.email], cc=cc)
        email.attach(f"fiche_palanquee_{palanquee.seance.date}_{palanquee.encadrant.nom_complet}.pdf", buffer.read(), 'application/pdf')
        email.attach_alternative(body_html, "text/html")
        if os.path.exists(signature_img_path):
            with open(signature_img_path, 'rb') as img:
                mime_img = MIMEImage(img.read(), _subtype='png')
                mime_img.add_header('Content-ID', '<signature_mouss2>')
                mime_img.add_header('Content-Disposition', 'inline', filename='Signature_mouss2.png')
                email.attach(mime_img)
        try:
            email.send()
            nb_envoyes += 1
        except Exception as e:
            erreurs.append(f"{encadrant.nom_complet} : {str(e)}")
    if nb_envoyes:
        messages.success(request, f"{nb_envoyes} PDF envoyés aux encadrants.")
    if erreurs:
        messages.error(request, "Erreurs lors de l'envoi : " + ", ".join(erreurs))
    return redirect('seance_detail', pk=seance_id)

@login_required
def envoyer_mail_covoiturage(request, seance_id):
    seance = get_object_or_404(Seance, pk=seance_id)
    inscriptions = seance.inscriptions.select_related('personne')
    conducteurs = [ins for ins in inscriptions if ins.covoiturage == 'propose']
    passagers = [ins for ins in inscriptions if ins.covoiturage == 'besoin']
    if not conducteurs or not passagers:
        messages.warning(request, "Aucun conducteur ou aucun passager à notifier.")
        return redirect('seance_detail', pk=seance_id)
    tableau = "<table border='1' cellpadding='4' cellspacing='0'><tr><th>Nom</th><th>Prénom</th><th>Email</th><th>Lieu</th></tr>"
    for ins in conducteurs:
        tableau += f"<tr><td>{ins.personne.nom}</td><td>{ins.personne.prenom}</td><td>{ins.personne.email}</td><td>{ins.lieu_covoiturage or ''}</td></tr>"
    tableau += "</table>"
    cc = getattr(settings, 'EMAIL_CC_COVOIT', [])
    signature_html = get_signature_html()
    signature_img_path = os.path.join(settings.BASE_DIR, 'static', 'Signature_mouss2.png')
    nb_envoyes = 0
    erreurs = []
    for ins in passagers:
        if not ins.personne.email:
            continue
        subject = f"Covoiturage pour la séance du {seance.date.strftime('%d/%m/%Y')}"
        body_html = f"Bonjour {ins.personne.prenom},<br><br>Voici la liste des personnes qui proposent du covoiturage pour la séance du {seance.date.strftime('%d/%m/%Y')} :<br><br>{tableau}<br><br>Merci de contacter directement les conducteurs pour organiser ton déplacement.<br><br>Subaquatiquement," + signature_html
        email = EmailMultiAlternatives(subject, '', to=[ins.personne.email], cc=cc)
        email.attach_alternative(body_html, "text/html")
        if os.path.exists(signature_img_path):
            with open(signature_img_path, 'rb') as img:
                mime_img = MIMEImage(img.read(), _subtype='png')
                mime_img.add_header('Content-ID', '<signature_mouss2>')
                mime_img.add_header('Content-Disposition', 'inline', filename='Signature_mouss2.png')
                email.attach(mime_img)
        try:
            email.send()
            nb_envoyes += 1
        except Exception as e:
            erreurs.append(f"{ins.personne.nom} {ins.personne.prenom} : {str(e)}")
    if nb_envoyes:
        messages.success(request, f"{nb_envoyes} mails de covoiturage envoyés.")
    if erreurs:
        messages.error(request, "Erreurs lors de l'envoi : " + ", ".join(erreurs))
    return redirect('seance_detail', pk=seance_id)

@csrf_exempt
@require_POST
@login_required
def envoyer_mail_inscription(request):
    import json
    from email.mime.image import MIMEImage
    try:
        data = json.loads(request.body)
        email = data.get('email')
        if not email:
            return JsonResponse({'success': False, 'error': 'Adresse email manquante.'})
        url = request.build_absolute_uri('/adherents/inscription/2025-2026/')
        subject = "Finalisation de ton inscription - Aquadémie Paris Plongée"
        signature_html = get_signature_html()
        signature_img_path = os.path.join(settings.BASE_DIR, 'static', 'Signature_mouss2.png')
        body_html = f"""
Bonjour,<br><br>
Je constate que tu as procédé à ton inscription sur HelloAsso et que tu as omis de renseigner le formulaire de l'étape n°1.<br>
<span style='color:red;font-weight:bold;'>Pour le moment tu n'as pas la possibilité de t'inscrire aux fosses.</span><br><br>
Pour y remédier, il faut finaliser ton inscription et renseigner ce formulaire : <a href='{url}'>{url}</a><br><br>
Je t'en remercie par avance.<br><br>
Subaquatiquement,<br>
""" + signature_html
        cc = getattr(settings, 'EMAIL_CC_DEFAULT', [])
        email_obj = EmailMultiAlternatives(subject, '', to=[email], cc=cc)
        email_obj.attach_alternative(body_html, "text/html")
        if os.path.exists(signature_img_path):
            with open(signature_img_path, 'rb') as img:
                mime_img = MIMEImage(img.read(), _subtype='png')
                mime_img.add_header('Content-ID', '<signature_mouss2>')
                mime_img.add_header('Content-Disposition', 'inline', filename='Signature_mouss2.png')
                email_obj.attach(mime_img)
        email_obj.send()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def valider_caci(request, adherent_id):
    adherent = get_object_or_404(Adherent, pk=adherent_id)
    if request.method == 'POST':
        adherent.caci_valide = True
        adherent.save()
        messages.success(request, f"CACI validé pour {adherent.nom} {adherent.prenom}.")
    return redirect('dashboard')

@login_required
def envoyer_liens_evaluation_encadrants(request, seance_id):
    seance = get_object_or_404(Seance, pk=seance_id)
    nb_envoyes = 0
    erreurs = []
    for palanquee in seance.palanques.all():
        encadrant = palanquee.encadrant
        if not encadrant or not encadrant.email:
            erreurs.append(f"{palanquee.nom} : pas d'encadrant ou d'email")
            continue
        # Chercher un lien d'évaluation actif ou en créer un
        from .models import LienEvaluation
        lien = LienEvaluation.objects.filter(palanquee=palanquee, date_expiration__gte=timezone.now()).last()
        if not lien:
            # Créer un nouveau lien
            import uuid
            from datetime import timedelta
            lien = LienEvaluation.objects.create(
                palanquee=palanquee,
                token=uuid.uuid4(),
                date_expiration=seance.date + timedelta(days=2)
            )
        # Envoyer le mail
        ok, msg = envoyer_lien_evaluation(lien, request)
        if ok:
            nb_envoyes += 1
        else:
            erreurs.append(f"{palanquee.nom} : {msg}")
    if nb_envoyes:
        messages.success(request, f"{nb_envoyes} mails de lien d'évaluation envoyés aux encadrants.")
    if erreurs:
        messages.error(request, "Erreurs lors de l'envoi : " + ", ".join(erreurs))
    return redirect('seance_detail', pk=seance_id)

def copier_caci(request, adherent_id):
    adherent = get_object_or_404(Adherent, pk=adherent_id)
    if not adherent.caci_fichier:
        messages.error(request, "Aucun fichier CACI à copier pour cet adhérent.")
        if request.GET.get('from') == 'dashboard':
            return redirect('dashboard')
        return redirect('adherent_list')
    chemin_sftp = getattr(settings, 'CHEMIN_SFTP', None)
    if not chemin_sftp:
        messages.error(request, "CHEMIN_SFTP non configuré dans les settings.")
        if request.GET.get('from') == 'dashboard':
            return redirect('dashboard')
        return redirect('adherent_list')
    nom = adherent.nom.strip().replace(' ', '_')
    prenom = adherent.prenom.strip().replace(' ', '_')
    ext = os.path.splitext(adherent.caci_fichier.name)[1]
    date_caci = ''
    if adherent.date_delivrance_caci:
        date_caci = adherent.date_delivrance_caci.strftime('%Y-%m-%d')
    else:
        date_caci = 'sansdate'
    nouveau_nom = f"CACI_{nom}_{prenom}-{date_caci}{ext}"
    source_path = adherent.caci_fichier.path
    dest_path = os.path.join(chemin_sftp, nouveau_nom)
    try:
        shutil.copy2(source_path, dest_path)
        messages.success(request, f"Fichier CACI copié sous {nouveau_nom}.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la copie : {e}")
    if request.GET.get('from') == 'dashboard':
        return redirect('dashboard')
    return redirect('adherent_list')

def copier_tous_caci(request):
    from .models import Adherent
    chemin_sftp = getattr(settings, 'CHEMIN_SFTP', None)
    if not chemin_sftp:
        messages.error(request, "CHEMIN_SFTP non configuré dans les settings.")
        return redirect('adherent_list')
    adherents = Adherent.objects.filter(caci_fichier__isnull=False).exclude(caci_fichier='')
    nb_copies = 0
    erreurs = []
    for adherent in adherents:
        nom = adherent.nom.strip().replace(' ', '_')
        prenom = adherent.prenom.strip().replace(' ', '_')
        ext = os.path.splitext(adherent.caci_fichier.name)[1]
        date_caci = ''
        if adherent.date_delivrance_caci:
            date_caci = adherent.date_delivrance_caci.strftime('%Y-%m-%d')
        else:
            date_caci = 'sansdate'
        nouveau_nom = f"CACI_{nom}_{prenom}-{date_caci}{ext}"
        source_path = adherent.caci_fichier.path
        dest_path = os.path.join(chemin_sftp, nouveau_nom)
        try:
            shutil.copy2(source_path, dest_path)
            nb_copies += 1
        except Exception as e:
            erreurs.append(f"{adherent.nom} {adherent.prenom} : {e}")
    if nb_copies:
        messages.success(request, f"{nb_copies} fichier(s) CACI copié(s) dans {chemin_sftp}.")
    if erreurs:
        messages.error(request, "Erreurs : " + ", ".join(erreurs))
    return redirect('adherent_list')

def creer_compte_adherent(request, adherent_id):
    from gestion.models import Adherent
    adherent = Adherent.objects.get(pk=adherent_id)
    if adherent.user:
        messages.warning(request, "Un compte est déjà associé à cet adhérent.")
        return redirect('adherent_list')
    # Créer le User
    email = adherent.email
    username = email
    first_name = adherent.prenom
    last_name = adherent.nom
    # S'assurer que le username est unique
    UserModel = get_user_model()
    if UserModel.objects.filter(username=username).exists():
        messages.error(request, "Un utilisateur avec cet email existe déjà.")
        return redirect('adherent_list')
    user = UserModel.objects.create_user(
        username=username,
        email=email,
        first_name=first_name,
        last_name=last_name,
        is_active=True  # Compte activé dès la création
    )
    # Associer à l'adhérent
    adherent.user = user
    adherent.save()
    # Déterminer le groupe
    statut = adherent.statut
    if statut == 'eleve':
        group_name = 'eleve'
    elif statut == 'encadrant':
        group_name = 'encadrant'
    else:
        group_name = 'admin'
    group, created = Group.objects.get_or_create(name=group_name)
    user.groups.clear()
    user.groups.add(group)
    # Générer le lien d'activation
    current_site = get_current_site(request)
    subject = "Activation de votre compte Aquadémie Paris Plongée"
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    activation_url = f"https://{current_site.domain}{reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})}"
    message = render_to_string('registration/account_activation_email.html', {
        'user': user,
        'activation_url': activation_url,
        'site_name': current_site.name,
        'domain': current_site.domain,
    })
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.email])
    messages.success(request, f"Compte utilisateur créé et mail d'activation envoyé à {user.email}.")
    return redirect('adherent_list')

@staff_member_required
def exporter_inscrits_seance_excel(request, seance_id):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from django.http import HttpResponse
    from gestion.models import Seance, InscriptionSeance

    seance = get_object_or_404(Seance, pk=seance_id)
    inscriptions = seance.inscriptions.select_related('personne').all()

    # Séparation
    eleves = [i.personne for i in inscriptions if i.personne.statut == 'eleve']
    encadrants = [i.personne for i in inscriptions if i.personne.statut == 'encadrant']
    covoiturage = [i for i in inscriptions if i.covoiturage in ['propose', 'besoin']]

    # Colonnes communes
    colonnes = [
        'Nom Prénom', 'Date de naissance', 'Téléphone', 'Email', 'Adresse', 'CP', 'Ville',
        'Date de délivrance CACI', 'Numéro licence', 'Niveau', 'Statut', 'Adhérent'
    ]

    wb = openpyxl.Workbook()
    ws_eleves = wb.active
    ws_eleves.title = 'Élèves'
    ws_encadrants = wb.create_sheet('Encadrants')
    ws_covoit = wb.create_sheet('Covoiturage')

    # Helper pour remplir
    def ligne_adherent(a):
        return [
            f"{a.nom} {a.prenom}",
            a.date_naissance.strftime('%d/%m/%Y') if a.date_naissance else '',
            a.telephone,
            a.email,
            a.adresse,
            a.code_postal,
            a.ville,
            a.date_delivrance_caci.strftime('%d/%m/%Y') if a.date_delivrance_caci else '',
            a.numero_licence,
            a.get_niveau_display() if hasattr(a, 'get_niveau_display') else a.niveau,
            a.get_statut_display() if hasattr(a, 'get_statut_display') else a.statut,
            'Oui' if getattr(a, 'type_personne', 'adherent') == 'adherent' else 'Non',
        ]

    # Élèves
    ws_eleves.append(colonnes)
    for a in eleves:
        ws_eleves.append(ligne_adherent(a))

    # Encadrants
    ws_encadrants.append(colonnes)
    for a in encadrants:
        ws_encadrants.append(ligne_adherent(a))

    # Covoiturage
    colonnes_covoit = colonnes + ['Type covoiturage', 'Lieu covoiturage']
    ws_covoit.append(colonnes_covoit)
    for ins in covoiturage:
        a = ins.personne
        row = ligne_adherent(a)
        row += [dict(InscriptionSeance.COVOITURAGE_CHOICES).get(ins.covoiturage, ''), ins.lieu_covoiturage or '']
        ws_covoit.append(row)

    # Largeur colonnes
    for ws in [ws_eleves, ws_encadrants, ws_covoit]:
        for i, col in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    # Réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nom_fichier = f"inscrits_seance_{seance.date.strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    wb.save(response)
    return response

def affecter_section_masse(request):
    from django.contrib import messages
    from django.shortcuts import redirect
    # On propose tous les adhérents (club ou non), statut élève, sans section
    adherents_sans_section = Adherent.objects.filter(statut='eleve').filter(sections=None)
    if request.method == 'POST':
        form = AffectationSectionMasseForm(request.POST, adherents_queryset=adherents_sans_section)
        if form.is_valid():
            section = form.cleaned_data['section']
            adherents = form.cleaned_data['adherents']
            for adherent in adherents:
                adherent.sections.add(section)
            messages.success(request, f"Section '{section}' affectée à {adherents.count()} adhérent(s).")
            return redirect('adherent_list')
    else:
        form = AffectationSectionMasseForm(adherents_queryset=adherents_sans_section)
    return render(request, 'gestion/affecter_section_masse.html', {'form': form, 'adherents': adherents_sans_section})

@login_required
@group_required('admin')
def evaluations_list(request):
    from gestion.models import Palanquee, EvaluationExercice, Adherent
    palanquees_evaluees_ids = (
        EvaluationExercice.objects.values_list('palanquee_id', flat=True).distinct()
    )
    palanquees = Palanquee.objects.filter(id__in=palanquees_evaluees_ids, encadrant__isnull=False)

    # Filtres
    date_seance = request.GET.get('date_seance')
    encadrant_id = request.GET.get('encadrant')
    eleve_id = request.GET.get('eleve')
    if date_seance:
        palanquees = palanquees.filter(seance__date=date_seance)
    if encadrant_id:
        palanquees = palanquees.filter(encadrant_id=encadrant_id)
    if eleve_id:
        palanquees = palanquees.filter(eleves__id=eleve_id)
    palanquees = palanquees.select_related('seance', 'encadrant').prefetch_related('eleves').order_by('-seance__date', 'encadrant__nom').distinct()

    # Pour les filtres dropdown
    encadrants = Adherent.objects.filter(statut='encadrant').order_by('nom', 'prenom')
    eleves = Adherent.objects.filter(statut='eleve').order_by('nom', 'prenom')
    dates = Palanquee.objects.filter(id__in=palanquees_evaluees_ids).values_list('seance__date', flat=True).distinct().order_by('-seance__date')

    context = {
        'palanquees': palanquees,
        'encadrants': encadrants,
        'eleves': eleves,
        'dates': dates,
    }
    return render(request, 'gestion/evaluations_list.html', context)

@require_GET
@staff_member_required
def api_modele_mail(request, modele_id):
    modele = get_object_or_404(ModeleMailSeance, pk=modele_id)
    return JsonResponse({
        'objet': modele.objet,
        'contenu': modele.contenu,
    })


@staff_member_required
def supprimer_modele_mail_adherents(request, modele_id):
    modele = get_object_or_404(ModeleMailAdherents, pk=modele_id)
    modele.delete()
    from django.contrib import messages
    messages.success(request, "Modèle supprimé avec succès.")
    return redirect(request.META.get('HTTP_REFERER', 'adherents_communiquer'))

@staff_member_required
def supprimer_modele_mail(request, modele_id):
    modele = get_object_or_404(ModeleMailSeance, pk=modele_id)
    modele.delete()
    from django.contrib import messages
    messages.success(request, "Modèle supprimé avec succès.")
    return redirect(request.META.get('HTTP_REFERER', 'seance_list'))

@staff_member_required
def envoyer_pdf_palanquee_encadrant(request, palanquee_id):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from io import BytesIO
    import os
    from django.conf import settings
    from .utils import get_signature_html
    from django.core.mail import EmailMultiAlternatives
    from django.template.loader import render_to_string
    from .models import Palanquee

    palanquee = get_object_or_404(Palanquee, pk=palanquee_id)
    encadrant = palanquee.encadrant
    seance = palanquee.seance
    if not encadrant or not encadrant.email:
        messages.error(request, "Aucun encadrant ou email pour cette palanquée.")
        return redirect('seance_detail', pk=seance.pk)
    # Générer le PDF en mémoire
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=12, spaceBefore=20
    )
    normal_style = styles['Normal']
    encadrant_nom = palanquee.encadrant.nom_complet if palanquee.encadrant else "-"
    elements.append(Paragraph(f"Palanquée : {encadrant_nom}", title_style))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Informations générales", heading_style))
    elements.append(Paragraph(f"<b>Date :</b> {palanquee.seance.date.strftime('%d/%m/%Y')}", normal_style))
    elements.append(Paragraph(f"<b>Lieu :</b> {palanquee.seance.lieu}", normal_style))
    elements.append(Paragraph(f"<b>Encadrant :</b> {encadrant_nom}", normal_style))
    elements.append(Paragraph(f"<b>Section :</b> {palanquee.section.get_nom_display()}", normal_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Élèves", heading_style))
    eleves_list = [eleve.nom_complet for eleve in palanquee.eleves.all()]
    elements.append(Paragraph(f"<b>Participants :</b> {', '.join(eleves_list)}", normal_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Exercices prévus", heading_style))
    exercices_list = [ex.nom for ex in palanquee.exercices_prevus.all()]
    for i, exercice in enumerate(exercices_list, 1):
        elements.append(Paragraph(f"{i}. {exercice}", normal_style))
    elements.append(Spacer(1, 12))
    if palanquee.precision_exercices:
        elements.append(Paragraph("Nota", heading_style))
        elements.append(Paragraph(palanquee.precision_exercices, normal_style))
    doc.build(elements)
    buffer.seek(0)
    # Préparer et envoyer le mail
    signature_html = get_signature_html()
    signature_img_path = os.path.join(settings.BASE_DIR, 'static', 'Signature_mouss.png')
    subject = f"Fiche palanquée - {palanquee.nom} ({seance.date})"
    body = render_to_string('gestion/email_pdf_palanquee.txt', {'palanquee': palanquee, 'seance': seance})
    body_html = f"<p>{body.replace(chr(10), '<br>')}</p>{signature_html}"
    email = EmailMultiAlternatives(subject, body, to=[encadrant.email])
    email.attach(f"fiche_palanquee_{palanquee.seance.date}_{palanquee.encadrant.nom_complet}.pdf", buffer.read(), 'application/pdf')
    email.attach_alternative(body_html, "text/html")
    if os.path.exists(signature_img_path):
        with open(signature_img_path, 'rb') as img:
            mime_img = MIMEImage(img.read(), _subtype='png')
            mime_img.add_header('Content-ID', '<signature_mouss2>')
            mime_img.add_header('Content-Disposition', 'inline', filename='Signature_mouss2.png')
            email.attach(mime_img)
    try:
        email.send()
        messages.success(request, f"PDF envoyé à l'encadrant {encadrant.nom_complet}.")
    except Exception as e:
        messages.error(request, f"Erreur lors de l'envoi : {str(e)}")
    return redirect('seance_detail', pk=seance.pk)

@method_decorator(staff_member_required, name='dispatch')
class CommunicationAdherentsView(LoginRequiredMixin, View):
    def get(self, request):
        adherents_qs = Adherent.objects.filter(actif=True).order_by('nom', 'prenom')
        adherents_choices = [
            (str(a.id), f"{a.nom.upper()} {a.prenom.capitalize()} ({a.email})")
            for a in adherents_qs
        ]
        form = CommunicationAdherentsForm(inscrits_choices=adherents_choices)
        modeles = ModeleMailAdherents.objects.all()
        historique = HistoriqueMailAdherents.objects.order_by('-date_envoi')[:20]
        return render(request, 'gestion/communication_adherents.html', {
            'form': form,
            'adherents': adherents_qs,
            'modeles': modeles,
            'historique': historique,
        })

    def post(self, request):
        adherents_qs = Adherent.objects.filter(type_personne='adherent').order_by('nom', 'prenom')
        adherents_choices = [
            (str(a.id), f"{a.nom.upper()} {a.prenom.capitalize()} ({a.email})")
            for a in adherents_qs
        ]
        form = CommunicationAdherentsForm(request.POST, request.FILES, inscrits_choices=adherents_choices)
        modeles = ModeleMailAdherents.objects.all()
        historique = HistoriqueMailAdherents.objects.all()[:20]

        # Suppression d'un modèle
        if 'supprimer_modele' in request.POST:
            modele_id = request.POST.get('modele')
            if modele_id:
                ModeleMailAdherents.objects.filter(id=modele_id).delete()
                messages.success(request, "Modèle supprimé avec succès.")
                return redirect('adherents_communiquer')

        # Enregistrement d'un modèle
        if 'enregistrer_modele' in request.POST:
            nom = request.POST.get('nom_modele', '').strip()
            objet = request.POST.get('objet', '').strip()
            contenu = request.POST.get('contenu', '').strip()
            if nom:
                ModeleMailAdherents.objects.create(
                    nom=nom,
                    objet=objet,
                    contenu=contenu,
                    auteur=request.user
                )
                messages.success(request, "Modèle enregistré avec succès.")
                return redirect('adherents_communiquer')
            else:
                messages.error(request, "Veuillez donner un nom au modèle.")
                return render(request, 'gestion/communication_adherents.html', {
                    'form': form,
                    'adherents': adherents_qs,
                    'modeles': modeles,
                    'historique': historique,
                })

        # Envoi du mail
        if form.is_valid():
            destinataires = []
            
            ids = form.cleaned_data['inscrits_choisis']
            destinataires = list(adherents_qs.filter(id__in=ids).values_list('email', flat=True))
            destinataires = [email.strip() for email in destinataires if email and email.strip()]
            destinataires = list(set(destinataires))
            if not destinataires:
                messages.error(request, "Aucun destinataire sélectionné.")
                return render(request, 'gestion/communication_adherents.html', {
                    'form': form,
                    'adherents': adherents_qs,
                    'modeles': modeles,
                    'historique': historique,
                })
            fichiers = request.FILES.getlist('fichiers')
            if len(fichiers) > 5:
                messages.error(request, "Vous ne pouvez joindre que 5 fichiers maximum.")
                return render(request, 'gestion/communication_adherents.html', {
                    'form': form,
                    'adherents': adherents_qs,
                    'modeles': modeles,
                    'historique': historique,
                })
            for f in fichiers:
                if f.size > 5*1024*1024:
                    messages.error(request, f"Le fichier {f.name} dépasse la taille maximale de 5Mo.")
                    return render(request, 'gestion/communication_adherents.html', {
                        'form': form,
                        'adherents': adherents_qs,
                        'modeles': modeles,
                        'historique': historique,
                    })
            from django.core.mail import EmailMessage
            from django.conf import settings
            # Envoi par lots de 10 avec pause
            batch_size = 10
            for i in range(0, len(destinataires), batch_size):
                batch = destinataires[i:i+batch_size]
                email = EmailMessage(
                    subject=form.cleaned_data['objet'],
                    body=form.cleaned_data['contenu'],
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    bcc=batch,
                )
                email.content_subtype = "html"
                for f in fichiers:
                    email.attach(f.name, f.read(), f.content_type)
                email.send()
                if (i + batch_size) < len(destinataires):
                    time.sleep(3)
            # Historique
            HistoriqueMailAdherents.objects.create(
                objet=form.cleaned_data['objet'],
                contenu=form.cleaned_data['contenu'],
                destinataires=",".join(destinataires),
                fichiers=",".join([f.name for f in fichiers]),
                auteur=request.user
            )
            messages.success(request, "Mail envoyé avec succès à :<br>" + "<br>".join(destinataires))
            return redirect('adherents_communiquer')
        else:
            return render(request, 'gestion/communication_adherents.html', {
                'form': form,
                'adherents': adherents_qs,
                'modeles': modeles,
                'historique': historique,
            })

def api_modele_mail_adherents(request, modele_id):
    modele = ModeleMailAdherents.objects.filter(id=modele_id).first()
    if not modele:
        return JsonResponse({'error': 'Modèle introuvable'}, status=404)
    return JsonResponse({
        'objet': modele.objet,
        'contenu': modele.contenu,
    })

@require_POST
def adherent_desactiver(request, pk):
    adherent = get_object_or_404(Adherent, pk=pk)
    adherent.actif = False
    adherent.save()
    return redirect('adherent_list')

@require_POST
def supprimer_historique_mail_adherents(request, mail_id):
    HistoriqueMailAdherents.objects.filter(id=mail_id).delete()
    messages.success(request, "Entrée d'historique supprimée.")
    return redirect('adherents_communiquer')

